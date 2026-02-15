#coding: utf-8
from decouple import Config, RepositoryEnv
from openpyxl import load_workbook
from celery.execute import send_task
import os, arrow
import logging
from tqdm import tqdm
from datetime import datetime, timedelta
from django.conf import settings
from OptsIO.io_serial import moneyfmt, format_codigo_barra, mes_palabra, url_viewfull
from Sifen.models import Business, DocumentHeader, DocumentDetail, Etimbrado
import pandas as pd
from fpdf import FPDF
import re

class EKude:
    def __init__(self, ruc: str):
        self.bobj = Business.objects.get(ruc=ruc)
        self.BASE_APP  = f'{settings.BASE_DIR}/Sifen'
        self.header_image = f'{self.BASE_APP}/assets/logo.png'
        self.header_image_x = 1.5
        self.header_image_y = 0.2
        self.header_image_w = 3.4
        self.header_image_h = 1.0
        self.mnpps = {'places':0, 'sep':'.'}
        self.paper = 'legal'
        self.coords = [0.7, 11.3, 22.1]

    def set_full_page(self):
        self.coords = [1.8]
        return {'exitos': 'Hecho'}
    
    def set_header_full(self, epdf, headerobj: DocumentHeader):
        #General Invoice info
        epdf.set_font('Arial',style='',size=7)
        doc_fecha = headerobj.doc_fecha
        impreso_sucursal = headerobj.doc_establecimiento
        impreso_caja = headerobj.doc_expedicion
        doc_numero = headerobj.doc_numero
        full_doc_numero = '{0:}-{1:}-{2: >7} '.format(str(impreso_sucursal).zfill(3),
                                                str(impreso_caja).zfill(3),
                                                str(doc_numero).zfill(7))
        for idx, coorde in enumerate(self.coords):
            epdf.set_font('Arial',style='B',size=6)
            epdf.set_y(coorde)
            epdf.set_x(9)
            label = 'KuDE de Factura Electronica'
            if headerobj.doc_tipo == 'NC':
                label = 'KuDE de Nota de Credito Electronica'
            if headerobj.doc_tipo == 'ND':
                label = 'KuDE de Nota de Debito Electronica'
            if headerobj.doc_tipo == 'AF':
                label = 'KuDE de Auto Factura Electronica'
            epdf.cell(0, 0,label)
            #epdf.set_y(coorde - 0.6)
            #Center title
            #Left beginning of the customer and invoice informacion
            epdf.set_y(coorde + 0.2)
            epdf.set_x(17.1)
            epdf.cell(0, 0,'RUC:')
            epdf.set_y(coorde + 0.5)
            epdf.set_x(16.4)
            epdf.cell(0, 0,'TIMBRADO:')
            epdf.set_y(coorde + 0.8)
            epdf.set_x(15.5)
            epdf.cell(0, 0,'INICIO DE VIGENCIA:')
            epdf.set_y(coorde + 1.1)
            epdf.set_x(15)
            if headerobj.doc_tipo in ['FE', 'FL', 'MI']:
                epdf.cell(0, 0,u'FACTURA ELECTRÓNICA:')
            if headerobj.doc_tipo in ['AF']:
                epdf.set_x(14.4)
                epdf.cell(0, 0,u'AUTOFACTURA ELECTRÓNICA:')
            if headerobj.doc_tipo == 'NC':
                epdf.set_x(14.4)
                epdf.cell(0, 0,u'NOTA CREDITO ELECTRÓNICA:')
            if headerobj.doc_tipo == 'ND':
                epdf.set_x(14.5)
                epdf.cell(0, 0,u'NOTA DEBITO ELECTRÓNICA:')
            epdf.set_font('Arial',style='',size=7.4)
            epdf.set_y(coorde + 1.1)
            epdf.set_x(17.8)
            if headerobj.doc_tipo in ['NC', 'ND', 'AF']:
                epdf.set_x(17.8)
            epdf.cell(0, 0,full_doc_numero)
            epdf.set_y(coorde + 0.2)
            epdf.set_x(17.8)
            epdf.cell(0, 0,'{}-{}'.format(self.bobj.ruc, self.bobj.ruc_dv))
            epdf.set_y(coorde + 0.5)
            epdf.set_x(17.8)
            epdf.cell(0, 0,'{}'.format(headerobj.ek_timbrado))
            epdf.set_y(coorde + 0.8)
            epdf.set_x(17.8)
            epdf.cell(0, 0,'{}'.format(headerobj.ek_timbrado_vigencia.strftime('%d/%m/%Y')))
            #Client data
            epdf.set_font('Arial',style='B',size=6.7)
            epdf.set_y(coorde+1.7)
            epdf.set_x(1)
            epdf.cell(0,0, 'Fecha y hora de emision:')
            epdf.set_x(16)
            if headerobj.doc_tipo in ['FE', 'FL']:
                epdf.cell(0,0, 'Vencimiento:')
            if headerobj.doc_tipo in ['AF']:
                epdf.set_x(14.5)
                epdf.cell(0,0, 'Tipo constancia:')
                epdf.set_font('Arial',style='',size=6.5)
                epdf.set_x(16.5)
                epdf.cell(0,0, headerobj.doc_relacion_tipo)
            epdf.set_font('Arial',style='B',size=6.7)
            epdf.set_y(coorde+2.1)
            epdf.set_x(1)
            if headerobj.doc_tipo in ['AF']:
                epdf.cell(0,0, 'RUC Comprador:')
            else:
                epdf.cell(0,0, 'Ruc/Documento Identidad:')
            epdf.set_y(coorde+2.1)
            if headerobj.doc_tipo in ['AF']:
                epdf.set_x(9.3)
                epdf.cell(0,0, 'CI Vendedor:')
            else:
                epdf.set_x(10.1)
                #epdf.cell(0,0, 'Codigo:')
            epdf.set_y(coorde+2.1)
            #if headerobj.doc_tipo in ['FE', 'FL']:
                # if headerobj.repartoobj:
                #     epdf.set_x(12.9)
                #     epdf.cell(0,0, 'Camion:')
            epdf.set_x(16)
            if headerobj.doc_tipo in ['FE', 'FL', 'MI']:
                epdf.cell(0,0, 'Tipo transaccion:')
            if headerobj.doc_tipo in ['NC', 'ND']:
                epdf.set_y(coorde+1.7)
                epdf.set_x(12.8)
                epdf.cell(0,0, 'Documento Asociado:')
            epdf.set_y(coorde+2.5)
            epdf.set_x(1)
            if headerobj.doc_tipo in ['AF']:
                epdf.cell(0,0, 'Nombre Comprador:')
            else:
                epdf.cell(0,0, 'Nombre o Razon Social:')
            if headerobj.doc_tipo in ['AF']:
                epdf.set_x(8.6)
                epdf.cell(0,0, 'Nombre Vendedor:')
            else:
                epdf.set_x(10.1)
                epdf.cell(0,0, 'Telef:')
            # if headerobj.doc_tipo in ['FE', 'FL']:
            #     if headerobj.repartoobj:
            #         epdf.set_x(13)
            #         epdf.cell(0,0, 'Chofer:')
            if headerobj.doc_tipo in ['AF']:
                if headerobj.doc_relacion_establecimiento:
                    epdf.set_x(14.9)
                    epdf.cell(0,0, 'Nro Control:')
            else:
                epdf.set_x(16)
                epdf.cell(0,0, 'Moneda:')

            if headerobj.doc_relacion and headerobj.doc_tipo in ['NC', 'ND']:
                epdf.set_y(coorde+2.1)
                epdf.set_x(12.8)
                epdf.cell(0,0, 'Motivo emision:')
                epdf.set_x(16)
                epdf.cell(0,0, 'Tipo documento asociado:')

            if headerobj.doc_relacion and headerobj.doc_tipo in ['AF']:
                if headerobj.doc_relacion_timbrado:
                    epdf.set_y(coorde+2.1)
                    epdf.set_x(14.5)
                    epdf.cell(0,0, 'Nro Constancia:')
                    epdf.set_font('Arial',style='',size=6.5)
                    epdf.set_x(16.5)
                    epdf.cell(0,0, str(headerobj.doc_relacion_timbrado))
            epdf.set_font('Arial',style='B',size=6.7)
            epdf.set_y(coorde+2.9)
            if headerobj.doc_tipo in ['AF']:
                epdf.set_x(1)
                epdf.cell(0,0, 'Direccion Transacion:')
                epdf.set_x(8.6)
                epdf.cell(0,0, 'Direccion Vendedor:')
                if headerobj.doc_relacion_expedicion:
                    epdf.set_x(14.9)
                    epdf.cell(0,0, 'Nro retencion:')
            else:
                epdf.set_x(1)
                epdf.cell(0,0, 'Direccion:')
                epdf.set_x(10.1)
                epdf.cell(0,0, 'Correo:')
                epdf.set_x(16)
                epdf.cell(0,0, 'Condicion de venta:')

            if headerobj.doc_tipo in ['FE', 'FL', 'AF', 'MI']:
                epdf.set_y(coorde+3.2)
                epdf.set_x(1)
                epdf.cell(0,0, 'Obs:')
                epdf.set_font('Arial',style='',size=6.5)
                epdf.set_x(1.7)
                #epdf.cell(0,0, headerobj.observacion if headerobj.observacion not in ['ND', None] else '')
            if headerobj.doc_tipo in ['AF']:
                epdf.set_y(coorde+3.2)
                epdf.set_x(8.9)
                epdf.set_font('Arial',style='B',size=6.7)
                epdf.cell(0,0, 'Tipo Vendedor:')
                epdf.set_font('Arial',style='',size=6.5)
                epdf.set_x(11)
                epdf.cell(0,0, 'Extranjero' if headerobj.pdv_es_contribuyente else 'No contribuyente')
            if headerobj.doc_tipo in ['NC', 'ND']:
                epdf.set_y(coorde+3.2)
                epdf.set_x(1)
                epdf.cell(0,0, 'SNC:')
                epdf.set_font('Arial',style='',size=6.5)
                epdf.set_x(1.7)
                # epdf.cell(0,0, str(headerobj.solicitud_cliente))

            epdf.set_font('Arial',style='',size=6.5)
            epdf.set_y(coorde+1.7)
            epdf.set_x(3.9)
            epdf.cell(0,0, 'ASUNCION, %s DE %s DE %s.' % (doc_fecha.strftime('%d'),
                                                mes_palabra(int(doc_fecha.strftime('%m'))),
                                                doc_fecha.strftime('%Y')
                                            ))

            if headerobj.doc_tipo in ['FE', 'FL']:
                epdf.set_x(18)
                vencimiento_factura = headerobj.doc_vencimiento
                epdf.cell(0,0,vencimiento_factura.strftime('%d/%m/%Y'))

            if headerobj.doc_tipo in ['NC', 'ND']:
                epdf.set_x(15.4)
                epdf.set_font('Arial',style='',size=5.8)
                if headerobj.doc_relacion_cdc:
                    epdf.cell(0,0, headerobj.doc_relacion_cdc)
                epdf.set_font('Arial',style='',size=6.5)                
            epdf.set_y(coorde+2.1)
            if headerobj.doc_tipo in ['AF']:
                epdf.set_x(3.8)
            else:
                epdf.set_x(4.1)
            if headerobj.doc_tipo in ['AF']:
                epdf.cell(0,0, '{}-{}'.format(self.bobj.ruc, self.bobj.ruc_dv))
            else:
                epdf.cell(0,0, '{}-{}'.format(headerobj.pdv_ruc, headerobj.pdv_ruc_dv))
            epdf.set_x(11)
            if headerobj.doc_tipo in ['AF']:
                epdf.cell(0,0, '{}-{}'.format(headerobj.pdv_ruc, headerobj.pdv_ruc_dv))
            else:
                #epdf.cell(0,0, headerobj.pdv_clientecod)
                pass
            # if headerobj.doc_tipo in ['FE', 'FL']:
            #     if headerobj.repartoobj:
            #         epdf.set_x(13.9)
            #         rptobj = headerobj.repartoobj
            #         rpts = '{}[{}]'.format(rptobj.camion_chapa, rptobj.pk)
            #         epdf.cell(0,0, rpts)

            if headerobj.doc_tipo in ['NC', 'ND']:
                epdf.set_x(14.7)
                if headerobj.doc_op in ['RPT', 'RS']:
                    epdf.cell(0,0, u'Devolución')
                elif headerobj.doc_op in ['GA', 'G']:
                    epdf.cell(0,0, u'Bonificación')                    
                else:
                    epdf.cell(0,0, u'Descuento')
                epdf.set_x(19.1)
                if headerobj.doc_relacion:
                    epdf.cell(0, 0,headerobj.doc_relacion)

            epdf.set_x(18)
            if headerobj.doc_tipo in ['FE', 'FL','MI']:
                epdf.cell(0,0, headerobj.doc_tipo_ope_desc)

            epdf.set_y(coorde + 2.5)
            epdf.set_x(3.8)
            epdf.set_font('Arial',style='',size=5.8)
            if headerobj.doc_tipo in ['AF']:
                epdf.cell(0, 0,self.bobj.name)
            else:
                epdf.cell(0, 0,headerobj.pdv_nombrefactura[0:50])
            epdf.set_font('Arial',style='',size=6.5)
            if headerobj.doc_tipo in ['AF']:
                epdf.set_x(11)
                epdf.cell(0, 0,headerobj.pdv_nombrefactura)
            else:
                epdf.set_x(11)
                epdf.cell(0, 0,headerobj.pdv_telefono)
            # if headerobj.doc_tipo in ['FE', 'FL']:
            #     if headerobj.repartoobj:
            #         epdf.set_x(13.9)
            #         epdf.cell(0, 0,headerobj.repartoobj.chofer_titular)
            if headerobj.doc_tipo in ['AF']:
                if headerobj.doc_relacion_establecimiento:
                    epdf.set_x(16.4)
                    epdf.cell(0,0, str(headerobj.doc_relacion_establecimiento))
            else:
                epdf.set_x(18)
                if headerobj.doc_moneda == 'GS':
                    epdf.cell(0,0, 'Guarani')
                if headerobj.doc_moneda == 'USD':
                    epdf.cell(0,0, '(USD) Dolar')

            if headerobj.doc_tipo in ['FE', 'FL', 'NC', 'MI', 'ND']:
                epdf.set_y(coorde+2.9)
                epdf.set_x(2.2)
                epdf.cell(0,0, headerobj.pdv_direccion_entrega[0:60] if headerobj.pdv_direccion_entrega else '')
            if headerobj.doc_tipo in ['AF']:
                epdf.set_y(coorde+2.9)
                epdf.set_x(3.8)
                epdf.cell(0,0, self.bobj.direccion[0:60] if self.bobj.direccion else '')
                epdf.set_x(11)
                epdf.cell(0,0, headerobj.pdv_direccion_entrega[0:60] if headerobj.pdv_direccion_entrega else '')
            else:
                epdf.set_x(11)
                if headerobj.pdv_email:
                    epdf.cell(0,0, headerobj.pdv_email[0:40])
            if headerobj.doc_tipo in ['AF']:
                if headerobj.doc_relacion_expedicion:
                    epdf.set_x(16.4)
                    epdf.cell(0,0, str(headerobj.doc_relacion_expedicion))
            else:
                epdf.set_x(18.3)
                epdf.cell(0,0, headerobj.doc_cre_tipo)
                epdf.set_y(coorde+2.9)
            # if headerobj.repartoobj:
            #     epdf.set_font('Arial',style='B',size=6)
            #     epdf.set_x(14.2)
            #     epdf.cell(0, 0, 'Camion:')
            #     epdf.set_x(15.1)
            #     epdf.set_font('Arial',style='',size=6)
            #     epdf.cell(0, 0, '{}:{}'.format(headerobj.repartoobj_id, headerobj.repartoobj.chofer_titular))
            #epdf.set_x(19.3)
            #epdf.cell(0, 0, headerobj.pdv_clientecod)
        return epdf

    
    def define_lines_full(self, epdf,digital, empresa, headerobj: DocumentHeader):
        for idx, coorde in enumerate(self.coords):
            epdf.line(0.9, coorde+0.0, 8.9, coorde+0.0) #primera linea
            epdf.line(12.2, coorde+0.0, 20.6, coorde+0.0) #primera linea
            epdf.line(0.9, coorde+1.4, 20.6, coorde+1.4) # contiene que separa el timbrado
            if headerobj.doc_tipo not in ['AF']:
                epdf.line(14.2, coorde+3.1, 20.6, coorde+3.1) # titulos valor de venta encima de exenta, g5 y g10
            epdf.line(0.9, coorde+3.4, 20.6, coorde+3.4) # contiene la linea que separa las observacion, fecha, etc            
            epdf.line(0.9, coorde+3.8, 20.6, coorde+3.8) # titulos de la table, arti, cantidad, precio, exenta, etc
            if digital:
                #LEFT
                epdf.image(self.header_image, 
                            x=self.header_image_x, 
                            y=coorde+self.header_image_y, 
                            w=self.header_image_w, 
                            h=self.header_image_h, 
                            type='', 
                            link='')
                #CENTER
                epdf.set_y(coorde + 0.3)
                epdf.set_x(9)
                epdf.cell(0, 0, self.bobj.name)
                # epdf.set_y(coorde + 0.6)
                # epdf.set_x(7)
                #epdf.cell(0, 0,self.bobj.actividadecoobj.nombre_actividad[0:69])
                #epdf.set_font('Arial',style='',size=6.5)
                # epdf.cell(0, 0,self.bobj.actividadecoobj.
                #           nombre_actividad[0:69])
                epdf.set_y(coorde + 0.4)
                epdf.set_x(6.5)
                epdf.multi_cell(8, 0.30, self.bobj.actividadecoobj.nombre_actividad)
                epdf.set_font('Arial',style='',size=8)
                epdf.set_y(coorde + 0.9)
                epdf.set_x(6)
                #epdf.cell(0, 0,f'{self.bobj.direccion} {self.bobj.numero_casa}')
                epdf.set_y(coorde + 1.2)
                epdf.set_x(7.7)
                epdf.cell(0, 0,f'{self.bobj.ciudadobj.nombre_ciudad} {self.bobj.telefono}')

            #separadores dettales - inicio
            #Valor de venta title
            ax = 1.8
            epdf.line(ax, coorde+3.4, ax, coorde+20) # separador de columna articulo
            dx = 9.2
            if headerobj.doc_tipo not in ['AF']:
                epdf.line(dx, coorde+3.4, dx, coorde+20) # separador de columna descripcion
            umx = 10.2
            epdf.line(umx, coorde+3.4, umx, coorde+20) # separador de columna medida
            canx = 11.2
            epdf.line(canx, coorde+3.4, canx, coorde+20) # separador de columna cantidad
            desx = 12.7
            epdf.line(desx, coorde+3.4, desx, coorde+20) # separador de columna descuento
            pux = 14.2
            if headerobj.doc_tipo not in ['AF']:
                epdf.line(pux, coorde+3.1, pux, coorde+20) # separador de columna precio unitario
            exex = 16.2
            if headerobj.doc_tipo not in ['AF']:
                epdf.line(exex, coorde+3.4, exex, coorde+20) # separador de columna exentas
                ivax = 18.5
                epdf.line(ivax, coorde+3.4, ivax, coorde+20) # separador de columna iva 
            epdf.line(0.9, coorde+20, 20.6, coorde+20) # linea de fin de detalle - comienzo subtotales
            #epdf.line(0.9, coorde+8.4, 20.6,coorde+8.4) # linea de fin de subtotales

            ###############
            #Lineas contenedoras
            if headerobj.doc_tipo in ['AF']:
                epdf.line(0.9, coorde + 0.0, 0.9, coorde + 20.5)
                epdf.line(20.6, coorde + 0.0, 20.6, coorde + 20.5)
            else:
                epdf.line(0.9, coorde + 0.0, 0.9, coorde + 22)
                epdf.line(20.6, coorde + 0.0, 20.6, coorde + 22)
            #epdf.line(0.9, coorde + 10, 20.6, coorde + 10)
            ####################
            if idx == 0: tipo = 'ORIGINAL: Comprador'
            if idx == 1: tipo = 'DUPLICADO: Arch.Tributario'
            if idx == 2: tipo = 'TRIPLICADO: Contabilidad'
            #epdf.set_font('Arial', style='', size=5.1)
            epdf.set_y(coorde+10.1)
            # epdf.set_x(19)
            # epdf.cell(1,0, tipo, align='C')        
        return epdf 

    def set_titles_full(self, epdf, headerobj: DocumentHeader):
        #Write titles
        epdf.set_font('Arial',style='B',size=7)
        for idx, coorde in enumerate(self.coords):
            epdf.set_y(coorde+3.6)
            epdf.set_x(0.9)
            epdf.cell(0,0, 'Cod')
            epdf.set_x(4)
            epdf.cell(0,0, 'Descripcion')            
            epdf.set_x(9.2)
            if headerobj.doc_tipo in ['AF']:
                epdf.set_x(10.3)
            else:
                epdf.cell(0,0, 'Medida')                        
                epdf.set_x(10.4)            
            epdf.cell(0,0, 'Cant.')                                    
            epdf.set_x(11.3)
            epdf.cell(0,0, 'P. Unitario')
            epdf.set_x(12.8)
            if headerobj.doc_tipo not in ['AF']:
                epdf.cell(0,0, 'Descuento')
                epdf.set_x(14.6)
                epdf.cell(0,0, 'Exentas')
                epdf.set_x(17)
                epdf.cell(0,0, '%5')
                epdf.set_x(19)
                epdf.cell(0,0, '%10')
        epdf.set_font('Arial',style='B',size=7)
        for idx, coorde in enumerate(self.coords):
            epdf.set_y(coorde+3.3)
            if headerobj.doc_tipo == 'AF':
                epdf.set_y(coorde+3.6)
            epdf.set_x(15.8)
            epdf.cell(0,0, 'Valor de Venta')
        return epdf

    def set_detail_full(self, epdf, headerobj: DocumentHeader, digital):
        if headerobj.doc_moneda == 'GS':
            self.mnpps = {'places':0, 'sep':'.'}
        if headerobj.doc_moneda == 'USD':
            self.mnpps = {'places':2, 'sep':','}

        epdf.set_font('Arial',style='',size=6.5)
        attr_cant = 'cantidad'
        if headerobj.doc_op == 'RS':
            attr_cant = 'cantidad_devolucion'
        attr_precio = 'precio_unitario'
        attr_descuento = 'descuento'
        attr_exenta = 'exenta'
        attr_gravada_5 = 'gravada_5'
        attr_gravada_10 = 'gravada_10'
        for idx, coorde in enumerate(self.coords):
            start_point = 4
            for pidx, p in enumerate(DocumentDetail.objects.filter(documentheaderobj=headerobj,anulado=False).exclude(prod_cod=90000).order_by('pk')):
                articulo_descripcion = p.prod_descripcion
                if isinstance(articulo_descripcion, str):
                    articulo_descripcion = articulo_descripcion.encode('utf-8').decode('latin-1', errors='replace')
                epdf.set_y(coorde+start_point)
                if len(articulo_descripcion) >= 44:
                    if pidx > 0:
                        epdf.set_y((coorde+start_point)+0.4)

                cantidad = getattr(p, attr_cant)
                precio_unitario = getattr(p, attr_precio)
                exenta = getattr(p, attr_exenta)
                gravada_5 = getattr(p, attr_gravada_5)
                gravada_10 = getattr(p, attr_gravada_10)
                descuento = getattr(p, attr_descuento)
                if headerobj.doc_op == 'RS':
                    if exenta:
                        exenta = precio_unitario*cantidad
                    if gravada_5:
                        gravada_5 = precio_unitario*cantidad
                    if gravada_10:
                        gravada_10 = precio_unitario*cantidad
                if p.per_descuento == 100:
                    setattr(p, attr_exenta,0)
                    setattr(p, attr_gravada_5,0)
                    setattr(p, attr_gravada_10,0)                    
                    # p.exenta = 0
                    # p.gravada_5 = 0                    
                    # p.gravada_10 = 0
                # epdf.set_y(coorde+start_point)
                epdf.set_x(1)
                epdf.cell(1,0, str(p.prod_cod), align='L')
                epdf.set_x(1.7)
                setl = None
                setc = 0
                
                if len(articulo_descripcion) >= 44:
                    epdf.set_y((coorde+start_point)+0.2)
                    print(pidx)
                    if pidx == 0:
                        epdf.set_y((coorde+start_point)-0.1)                    
                    epdf.set_x(1.8)
                    setl = 45
                    setc = -6
                    epdf.multi_cell(7, 0.30, articulo_descripcion)
                    if pidx == 0:
                        epdf.set_y((coorde+start_point))
                    else:
                        epdf.set_y((coorde+start_point)+0.4)
                else:
                    epdf.cell(1,0, articulo_descripcion[0:setl], align='L')
                if headerobj.doc_tipo in ['FE', 'FL', 'NC', 'ND']:
                    epdf.set_x(8.2)
                    #epdf.cell(1,0, format_codigo_barra(artobj.get_codigobarra_fancy())[setc:], align='R')
                    epdf.set_x(9.4)
                    epdf.cell(1,0, p.prod_unidad_medida_desc, align='L')
                epdf.set_x(10.1)
                epdf.cell(1,0, str(cantidad), align='R')
                epdf.set_x(11.6)
                epdf.cell(1,0, moneyfmt(precio_unitario, **self.mnpps), align='R')
                if headerobj.doc_tipo not in ['AF']:
                    epdf.set_x(13.2)
                    epdf.cell(1,0, moneyfmt(descuento, **self.mnpps),align='R')
                    epdf.set_x(15.2)
                else:
                    epdf.set_x(16.3)
                epdf.cell(1,0, moneyfmt(exenta, **self.mnpps),align='R')
                if headerobj.doc_tipo not in ['AF']:
                    epdf.set_x(17.5)
                    epdf.cell(1,0, moneyfmt(gravada_5, **self.mnpps),align='R')
                    epdf.set_x(19.6)
                    epdf.cell(1,0, moneyfmt(gravada_10, **self.mnpps),align='R')
                start_point += 0.3
        return epdf

    def set_footer_full(self, epdf, headerobj: DocumentHeader):
        if headerobj.doc_moneda == 'GS':
            self.mnpps = {'places':0, 'sep':'.'}
        if headerobj.doc_moneda == 'USD':
            self.mnpps = {'places':2, 'sep':','}
        af_x = 22.2+0.5
        af_t = 22.5+0.5
        af_y = 23+0.5
        af_z = 23.5+0.5
        af_v = 24+0.5
        af_h = 24.4+0.5
        af_p = 24.8+0.5
        if headerobj.doc_tipo in ['AF']:
            af_x = 20.8
            af_t = 21.1
            af_y = 21.5
            af_z = 21.9
            af_v = 22.3
            af_h = 22.7
            af_p = 23.1

        epdf.set_font('Arial',style='',size=7)
        for idx, coorde in enumerate(self.coords):
            if headerobj.doc_tipo in ['AF']:
                epdf.line(0.9, coorde+20.5, 20.6, coorde+20.5) # linea de fin de detalle 
            else:
                epdf.line(0.9, coorde+20.5, 20.6, coorde+20.5) # linea de fin de detalle 
                epdf.line(0.9, coorde+21, 20.6, coorde+21) # linea de fin de detalle
                epdf.line(0.9, coorde+21.5, 20.6, coorde+21.5) # linea de fin de detalle
                epdf.line(0.9, coorde+22, 20.6, coorde+22) # linea de fin de detalle
            # epdf.line(0.9, coorde+10.6, 20.6, coorde+10.6) # linea de fin de detalle            
            #Titles of totals
            if headerobj.doc_tipo in ['AF']:
                epdf.set_font('Arial',style='B',size=6.7)
                epdf.set_y(coorde+20.3)
                epdf.set_x(10.6)
                epdf.cell(0,0, 'TOTAL A PAGAR:', align='L')
                epdf.set_y(coorde+20.3)
                epdf.set_x(16.5)
                epdf.cell(1,0, moneyfmt(headerobj.doc_exenta, **self.mnpps),align='R')
            else:
                epdf.set_font('Arial',style='B',size=6)
                epdf.set_y(coorde+20.3)
                epdf.cell(0,0, 'SUBTOTAL:', align='L')
                epdf.set_y(coorde+20.8)
                epdf.cell(0,0, 'TOTAL DE LA OPERACION:', align='L')            
                epdf.set_y(coorde+21.3)
                if headerobj.doc_moneda == 'GS':
                    epdf.cell(0,0, 'TOTAL EN GUARANIES:', align='L')
                if headerobj.doc_moneda == 'USD':
                    epdf.cell(0,0, 'TOTAL EN DOLARES AMERICANOS:', align='L')                    
                epdf.set_y(coorde+21.8)
                epdf.cell(0,0, 'LIQUIDACION IVA:', align='L')                        
                epdf.set_x(6)
                epdf.cell(0,0, '(%5)', align='L')
                epdf.set_x(9.5)
                epdf.cell(0,0, '(%10)', align='L')
                epdf.set_x(15)            
                epdf.cell(0,0, 'TOTAL IVA:', align='L')            
                #words
                epdf.set_y(coorde+22.3)
                if headerobj.doc_moneda == 'GS':
                    epdf.cell(0,0, 'TOTAL EN GUARANIES:', align='L')
                if headerobj.doc_moneda == 'USD':
                    epdf.cell(0,0, 'TOTAL EN DOLARES AMERICANOS:', align='L')
                #SET VALUES
                epdf.set_font('Arial',style='',size=6)
                epdf.set_y(coorde+20.3)
                epdf.set_x(15.1)
                epdf.cell(1,0, moneyfmt(headerobj.doc_exenta, **self.mnpps),align='R')            
                epdf.set_x(17.5)
                epdf.cell(1,0, moneyfmt(headerobj.get_total_gravada_5(), **self.mnpps),align='R')
                epdf.set_x(19.6)
                epdf.cell(1,0, moneyfmt(headerobj.get_total_gravada_10(), **self.mnpps),align='R')
                epdf.set_font('Arial',style='',size=7)
                epdf.set_y(coorde+20.8)
                epdf.set_x(19.6)
                epdf.cell(1,0, moneyfmt(headerobj.get_total_venta(), **self.mnpps), align='R')

                epdf.set_y(coorde+21.3)
                epdf.set_x(19.6)
                epdf.cell(1,0, moneyfmt(headerobj.get_total_venta(), **self.mnpps), align='R')

                epdf.set_y(coorde+21.8)
                epdf.set_x(8)
                epdf.cell(1,0, moneyfmt(headerobj.get_ivas_5_master(), **self.mnpps), align='R')
                epdf.set_x(13)
                epdf.cell(1,0, moneyfmt(headerobj.get_ivas_10_master(), **self.mnpps), align='R')            

                epdf.set_x(19.6)
                epdf.cell(1,0, moneyfmt(headerobj.get_ivas_master(), **self.mnpps), align='R')

                epdf.set_y(coorde+22.3)
                epdf.set_x(14)
                vword = headerobj.get_total_ventaword()
                if len(vword) >= 85:
                    epdf.set_x(16)
                epdf.cell(1,0, vword, align='R')

            #QR AND CDC
            epdf.set_font('Arial',style='',size=9)
            if headerobj.ek_qr_img:
                epdf.image(headerobj.ek_qr_img.path, 
                    x=1.5, 
                    y=coorde+af_x, #22.2+0.5, 
                    w=3, 
                    h=3, 
                    type='', 
                    link=''
                )
            epdf.set_y(coorde+af_t) #22.5+0.5
            epdf.set_x(5)
            epdf.cell(0,0, u'Consulte la validez de esta Factura Electrónica con el número de CDC impreso abajo en:')
            epdf.set_y(coorde+af_y) #23+0.5
            epdf.set_x(5)
            epdf.cell(0,0, 'https://ekuatia.set.gov.py/consultas')
            epdf.set_y(coorde+af_z) #23.5+0.5
            epdf.set_x(5)
            cdcf = ' '.join([headerobj.ek_cdc[i:i+4] for i in range(0, len(headerobj.ek_cdc), 4)])
            epdf.set_font('Arial',style='B',size=12)
            epdf.cell(0,0, 'CDC: {}'.format(cdcf))
            epdf.set_font('Arial',style='',size=9)
            epdf.set_y(coorde+af_v) #24+0.5
            epdf.set_x(5)
            epdf.cell(0,0,u'ESTE DOCUMENTO ES UNA REPRESENTACIÓN GRÁFICA DE UN DOCUMENTO ELECTRÓNICO (XML)')
            epdf.set_y(coorde+af_h) #24.4+0.5
            epdf.set_x(5)            
            epdf.cell(0,0,u'Si su documento electrónico presenta algún error, podrá solicitar la modificación dentro de las')
            epdf.set_y(coorde+af_p) #24.8+0.5
            epdf.set_x(5)                      
            epdf.cell(0,0,u'72 horas siguientes de la emisión de este comprobante.')
        return epdf

    def eprint_docs(self, *args, **kwargs):
        userobj = args[0]
        qdict = kwargs.get('query_dict')
        impresora = qdict.get('impresora')
        digital = qdict.get('digital')
        layout = qdict.get('layout')
        ncopies = qdict.get('ncopies', 1)
        set_name = qdict.get('set_name')
        peds = qdict.getlist('prof_number')
        rsp = []
        for p in peds:
            rsp.append(self.eprint_doc(userobj, query_dict={'prof_number': p, 
                                                            'layout': layout, 
                                                            'impresora': impresora, 
                                                            'digital':digital, 
                                                            'set_name': set_name, 
                                                            'ncopies': ncopies}))
        #bbs = set(map(lambda x: x.get('bloquepk'), rsp))
        # for bb in bbs:
        #     send_task('sales_man.tasks.notification_invoice', args=(bb, ),
        #             kwargs={})
        return {'exitos': 'Hecho', 'prints': rsp}

    def eprint_doc(self, *args, **kwargs):
        now = datetime.now()
        tnow = now.strftime('%Y%m%d')
        userobj = args[0]
        
        qdict = kwargs.get('query_dict')
        prof_number = qdict.get('prof_number')
        impresora = qdict.get('impresora')
        digital = qdict.get('digital')
        set_name = qdict.get('set_name')
        layout = qdict.get('layout')
        ncopies = qdict.get('ncopies', 1)
        # if copies:
        #     self.set_copies(int(copies))
        headerobj = DocumentHeader.objects.get(prof_number=prof_number)
        
        #Header data
        doc_numero = headerobj.doc_numero
        if not doc_numero: return {'error': 'El documento carece de numero'}
        doc_condicion_pago = headerobj.doc_cre_tipo
        doc_cre_plazo = headerobj.doc_cre_plazo
        doc_moneda = 'PYG'
        if headerobj.doc_moneda != 'GS':
            moneda = headerobj.doc_moneda
        doc_relacion = 0
        if headerobj.doc_loop_link:
            relacion = headerobj.doc_loop_link
        pdv_ruc = headerobj.pdv_ruc
        pdv_razon = headerobj.pdv_nombrefactura
        pdv_direccion = headerobj.pdv_direccion_entrega
        pdv_telefono = headerobj.pdv_telefono
        pdv_email = headerobj.pdv_email
        tipo_transaccion = headerobj.doc_tipo_ope_desc
        # obs = u"{} PED {}".format(headerobj.observacion, headerobj.prof_number).strip()
        impreso_sucursal = headerobj.doc_establecimiento
        impreso_caja = headerobj.doc_expedicion
        timbrado = headerobj.ek_timbrado
        timbrado_vigencia = headerobj.ek_timbrado_vigencia.strftime('%d/%m/%Y')
        timbrado_vencimiento = headerobj.ek_timbrado_vencimiento.strftime('%d/%m/%Y')
        #totales dineros en descuento PowerPack
        full_doc_numero = '{0:}-{1:}-{2: >7} '.format(str(impreso_sucursal).zfill(3),
                                                str(impreso_caja).zfill(3),
                                                str(doc_numero).zfill(7))
        camion = None                                                
        chofer = None
        rpt = 0

        pdf=FPDF('P','cm',self.paper)
        pdf.add_page()
        pdf.set_font('Arial',style='',size=8)
        pdf.set_line_width(0)
        pdf.set_draw_color(44, 46, 45)
        fftname = headerobj.ek_cdc
        if set_name == 'doc_numero':
            fftname = '{}_{}'.format(
                headerobj.doc_tipo,
                full_doc_numero.split('-')[-1].strip()
            )
        dfname = '{}_full_page'.format(fftname)
        self.set_full_page()
        pdf = self.define_lines_full(pdf, digital, self.bobj.abbr, headerobj)
        pdf = self.set_header_full(pdf, headerobj)
        pdf = self.set_titles_full(pdf, headerobj)
        pdf = self.set_detail_full(pdf, headerobj, digital)
        pdf = self.set_footer_full(pdf, headerobj)
        doc_dir = '{}/invoicing_files/{}'.format(self.BASE_APP, tnow)
        pdf_path = '{}/{}.pdf'.format(doc_dir,dfname)
        try:os.mkdir(doc_dir)
        except:pass
        pdf.output(pdf_path,'F')
        headerobj.ek_pdf_file = pdf_path
        headerobj.save()
        #TODO: See FL Printers
        # if impresora:
        #     cups_server = impretions.CupsPrinter()
        #     params_impresion = {'impresorapk': impresora,'ruta': pdf_path,'opciones': {}}
        #     for a in range(int(ncopies)):
        #         cups_server.imprimir_trabajo(**params_impresion)
        #     headerobj.documento_impreso = True
        #     headerobj.save()
        bloquepk = None
        if headerobj.doc_tipo == 'FE':
            bloquepk = None
        return  {'exitos': 'Hecho', 
                'pdf_file': pdf_path,
                #'location_file': url_viewfull('show_pdf_file', filename=pdf_path),
                'filename': dfname,
                'doc_numero': full_doc_numero.split('-')[-1].strip(),
                'bloquepk': bloquepk
            }

    def send_kude_console(self, *args, **kwargs):
        qdict = kwargs.get('query_dict')
        e_timbrado = qdict.get('e_timbrado')
        tipo = qdict.get('tipo')
        clientecod = qdict.get('clientecod')
        doc_numero = qdict.get('doc_numero')
        date_range = qdict.get('date_range')
        exclude_f = qdict.get('exclude_f')
        excf = {}
        if exclude_f:
            excf.update({'doc_fecha__range': exclude_f})
        kps = {
            'ek_estado':'Aprobado',
            'pedido_tipo':tipo,
            'e_timbrado':e_timbrado,
            'doc_entregado_por_gecos__isnull':True
        }
        if clientecod:
            kps['pdv_clientecod'] = clientecod
        if date_range:
            kps['doc_fecha__range'] = date_range
        if doc_numero:
            kps['doc_numero'] = doc_numero
        logging.info('Traer documentos con parametros {} exclusion {}'.format(kps, excf))
        prof_number = DocumentHeader.objects.filter(**kps)\
                              .values_list('prof_number', flat=True)\
                              .order_by('pdv_ruc', 'doc_numero')\
                              .exclude(**excf)
        rsp = []
        for idx, p in enumerate(tqdm(prof_number)):
            try:
                rsp.append(self.send_kude(*args, query_dict={'prof_number': p}))
            except Exception as e:
                print(e)
                pass
        if qdict.get('not_report'): return {'exitos': 'Hecho'}
        ipm = mng_rpt.RptManag()
        if not rsp: return {'exitos': 'Sin acciones que realizar'}
        label_d = 'FACTURA ELECTRONICA'
        if tipo == 'NC':
            label_d = 'NOTA DE CREDITO ELECTRONICA'
        if tipo == 'ND':
            label_d = 'NOTA DE DEBITO ELECTRONICA'
        if tipo == 'RE':
            label_d = 'NOTA DE REMISION ELECTRONICA'
        mme = list(set([ u'{}</li>'.format(f) for f in map(lambda x: x.get('error'), filter(lambda x: x.get('error'), rsp))]))
        HBODY = u"""
              <h1>Listado de operaciones de envio de {} a clientes</h1>
              <h2>NO FUE POSIBLE EL ENVIO DE LOS SIGUIENTES DOCUMENTOS</h2>
              <ul>{}</ul>
              <h2>ENVIOS EXITOSOS</h2>
              <ul>{}</ul>
        """.format(
            label_d,
            u'<li>'.join(mme),
            u'<li>'.join([ u'{}</li>'.format(f) for f in map(lambda x: x.get('exitos'), filter(lambda x: x.get('exitos'), rsp))])
        )
        dst = list(
                UserNotification.objects.filter(
                        operation='send_kude_console', 
                        anulado=False)\
                    .values_list('mail', flat=True)\
                    .distinct('mail')
        )
        if settings.DEBUG:
            dst = ['supervisor_red@rkf.com.py', 'desarrollo2@rkf.com.py', 'desarrollo3@rkf.com.py']
        ipm.static_report_mail(
                html=HBODY,
                modulo=u'[DOCUMENTO {}] Resumen de envios a clientes'.format(label_d),
                destinatarios=','.join(dst),
                esender='sifen@aconcagua.com.py',
                bcc = ['sifen_send@rkf.com.py']
        )
        return {'exitos': 'Hecho'}

    def send_kude(self, *args, **kwargs):
        qdict = kwargs.get('query_dict')
        prof_number = qdict.get('prof_number')
        self.eprint_doc(*args, 
                query_dict={'prof_number': prof_number, 
                            'layout': 'full_page', 
                            'digital': 1})
        headerobj = DocumentHeader.objects.get(prof_number=prof_number)

        full_doc_numero = '{0:}-{1:}-{2: >7} '.format(
            str(headerobj.doc_establecimiento).zfill(3),
            str(headerobj.doc_expedicion).zfill(3),
            str(headerobj.doc_numero).zfill(7)
        )        
        
        if not settings.DEBUG:
            if not headerobj.pdv_email or headerobj.pdv_email == 'ND':
                msge = u'El cliente {}[{}] No tiene correo'.format(
                    headerobj.pdv_nombrefantasia,
                    headerobj.pdv_clientecod,
                )
                return {'error': msge}
        label_d = u'Factura Electronica'
        if headerobj.doc_tipo == 'NC':
            label_d = u'Nota de credito Electronica'
        if headerobj.doc_tipo == 'ND':
            label_d = u'Nota de debito Electronica'
        if headerobj.doc_tipo == 'RE':
            label_d = u'Nota de remision Electronica'            
        TBODY=u"""
            <hr>
            {}
            Buenas Estimado Cliente {}.
            <br><br>
            Remitimos adjunto su {}.
            <br><br>
            Número {}: {}<br><br>
            Importe: {} {}<br><br>
            Link de Consulta :
            <a href="{}" target="_blank">{}</a>
            <br><br>
            Por favor no responda este e-mail.
            <br><br>
            <hr>
        """
        BODY = TBODY.format(
            '',
            headerobj.pdv_nombrefactura,
            label_d, label_d,
            full_doc_numero,
            moneyfmt(headerobj.get_total_venta(), **self.mnpps),
            headerobj.doc_moneda,
            headerobj.qr_link,
            headerobj.qr_link
        )
        ipm = mng_rpt.RptManag()

        # intccs = list(UserNotification.objects.filter(
        #                             operation='send_kude_failure', 
        #                             anulado=False)\
        #                     .values_list('mail', flat=True)\
        #                     .distinct('mail'))        
        dst = []
        if settings.DEBUG:
            dst = ['desarrollo2@rkf.com.py', 'supervisor_red@rkf.com.py', 'desarrollo3@rkf.com.py'
                ]
        else:
            if headerobj.pdv_email and headerobj.pdv_email != 'ND' and headerobj.pdv_email != '':
                dst = headerobj.pdv_email.split(',')
                dst = list(map(lambda x: x.strip().strip(','), dst))
        msgs = u'Documento: {} Tipo: {} Cliente: {}[{}] Receptores: {}'.format(
                                full_doc_numero, 
                                headerobj.doc_tipo, 
                                headerobj.pdv_nombrefantasia,
                                headerobj.pdv_clientecod, 
                                ','.join(dst)
                )
        logging.info(u'Enviando a correos {} cliente {}[{}] documento {}'.format(
            ','.join(dst), 
            headerobj.pdv_nombrefantasia, 
            headerobj.pdv_clientecod, 
            headerobj.pdf_file.split('/')[-1]
        ))
        #print(','.join(dst), headerobj.pdv_clientecod)
        ipm.static_report_mail(
                html=BODY,
                modulo=u'{} {} Cliente [{}]{} Emitida de {}'.format(label_d, 
                full_doc_numero.replace('-', ''), 
                headerobj.pdv_clientecod, 
                headerobj.pdv_nombrefactura,
                self.bobj.name
                ),
                destinatarios=','.join(dst),
                esender='supervisor_red@rkf.com.py' if settings.DEBUG else self.bobj.correo,
                filename=[headerobj.pdf_file, headerobj.xml_file],
                bcc = ['supervisor_red@rkf.com.py'] if settings.DEBUG else [self.bobj.correo],
                headers = {'Disposition-Notification-To': self.bobj.correo}
            )
        headerobj.doc_entregado = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        headerobj.doc_entregado_por_gecos = 'ENVIADO_CORREO'
        headerobj.save()
        return {'exitos': msgs}
    