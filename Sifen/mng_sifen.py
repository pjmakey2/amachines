import uuid, os
import requests
import pandas as pd
from zoneinfo import ZoneInfo
from collections import namedtuple, OrderedDict
from decimal import Decimal
import arrow, os, logging
from datetime import datetime, date, timedelta
from OptsIO.io_decorators import tracktask
from OptsIO.io_serial import IoS
from OptsIO.io_rpt import IoRpt
from OptsIO.io_json import to_json, from_json
from OptsIO import io_styles
from typing import Literal, Union
from django.contrib.auth.models import User
from django.http import QueryDict, HttpRequest
from django.forms import model_to_dict
from django.db.models import Sum
from django.core.files import File
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.conf import settings
from Sifen.models import DocumentHeader, DocumentDetail, Business, Etimbrado, Eestablecimiento, DocumentRecibo, DocumentReciboDetail, Departamentos, Distrito, Ciudades, Retencion, Cotizacion, Producto, Clientes
from Sifen import fl_sifen_conf, ekuatia_serials, e_kude, mng_gmdata
from OptsIO.models import UserProfile, UserBusiness
from Finance import f_calcs
from celery.execute import send_task
#from django.utils.html import strip_tags
from num2words import num2words
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class MSifen:
    def __init__(self, userobj=None, business=None):
        """
        Inicializa MSifen con el Business del usuario o uno proporcionado.

        Args:
            userobj: Usuario de Django (opcional)
            business: Objeto Business directo (opcional)
        """
        self.bsobj = None
        self.RUC = None
        self.timbradoobj = None

        # Prioridad: business directo > business del usuario > fallback fl_sifen_conf
        if business:
            self.bsobj = business
        elif userobj:
            self.bsobj = self._get_business_from_user(userobj)

        # Fallback a fl_sifen_conf si no hay business
        if not self.bsobj:
            try:
                self.RUC = fl_sifen_conf.RUC
                self.bsobj = Business.objects.get(ruc=self.RUC)
            except Business.DoesNotExist:
                # Si no existe, intentar obtener el primer Business disponible
                self.bsobj = Business.objects.first()

        if self.bsobj:
            self.RUC = self.bsobj.ruc
            try:
                self.timbradoobj = Etimbrado.objects.filter(ruc=self.RUC).first()
            except Etimbrado.DoesNotExist:
                self.timbradoobj = None

        self.asuzone = ZoneInfo('America/Asuncion')
        now = datetime.now(tz=self.asuzone)
        self.tnow = now.strftime('%Y-%m-%d')
        self.ttime = datetime.now(ZoneInfo('America/Asuncion')) - timedelta(minutes=2)

    def _get_business_from_user(self, userobj):
        """Obtiene el Business activo del usuario."""
        try:
            profile = UserProfile.objects.filter(username=userobj.username).first()
            if not profile:
                return None

            active_ub = UserBusiness.objects.filter(
                userprofileobj=profile,
                active=True
            ).select_related('businessobj').first()

            if active_ub and active_ub.businessobj:
                return active_ub.businessobj
        except Exception as e:
            logging.warning(f'Error obteniendo business del usuario: {e}')

        return None

    def send_invoice_from_ui(self, *args: list, **kwargs: dict) -> tuple:
        """
        Send invoice to client from UI
        """
        q: dict = kwargs.get('qdict', {})
        pk = q.get('id')
        docobj = DocumentHeader.objects.get(pk=pk)
        if docobj.ek_estado != 'Aprobado':
            return {'error': 'Solo se pueden enviar facturas aprobadas en la SIFEN'}, args, kwargs
        qq = QueryDict(mutable=True)
        qq.update({
            'docpk': pk,
            'dbcon': q.get('dbcon'),
            'force_send': True
        })
        return self.send_invoice(*args, qdict=qq)
        
    @tracktask
    def send_invoice(self, *args: list, **kwargs: dict) -> tuple:
        q: dict = kwargs.get('qdict', {})
        force_send = q.get('force_send', False)
        userobj = kwargs.get('userobj')
        if not userobj:
            userobj = User.objects.get(username='amadmin')
        docpk = q.get('docpk')
        dbcon = q.get('dbcon')
        from_console = q.get('from_console')
        docobj = DocumentHeader.objects.get(pk=docpk)
        logging.info(f'Enviando factura {docobj.doc_numero} al cliente {docobj.pdv_email}')
        tipo = 'Factura'
        if docobj.doc_tipo in ['NC']:
            tipo = 'Nota de credito'        
        if not force_send:
            if docobj.enviado_cliente  and from_console:
                return {'error': 'El correo ya fue enviado al cliente '}

        dattrs = {'media_path': True}

        # Get Business object data based on ek_bs_ruc
        bobj = Business.objects.get(ruc=docobj.ek_bs_ruc)
        bobj_dict = model_to_dict(bobj, exclude=['contribuyenteobj', 'actividadecoobj'])
        bobj_dict['contribuyente'] = bobj.contribuyenteobj.tipo
        bobj_dict['ciudad'] = bobj.ciudadobj.nombre_ciudad
        bobj_dict['denominacion'] = bobj.actividadecoobj.nombre_actividad
        # Convert ciudadobj FK to name string
        dattrs.update(bobj_dict)

        # Add logo from configuration
        #dattrs['sifen_logo'] = fl_sifen_conf.LOGO
        logging.info(f'Generando PDF para factura {docobj.doc_numero}')
        d_pdf = self.crear_documentheader_pdf(
            userobj=userobj,
            qdict={
                'dbcon': dbcon,
                'id': docpk,
                'dattrs': to_json(dattrs)
        })
        pdf_file = f'{settings.BASE_DIR}/{d_pdf.get("pdf_file")}'
        logging.info(f'PDF generado en {pdf_file}')
        xml_file = docobj.ek_xml_file_signed.path
        tp_l = {
            "tipo": tipo,
            "razon_social": docobj.pdv_nombrefactura,
            "ruc": docobj.pdv_ruc,
            "monto": docobj.get_total_venta(),
            "link": docobj.ek_qr_link,
            "numero": '{}-{}-{}'.format(
                str(docobj.doc_establecimiento).zfill(3),
                str(docobj.doc_expedicion).zfill(3),
                str(docobj.doc_numero).zfill(7),
            )
        }

        #html_c = render_to_string("Sifen/DocumentHeaderEmailUi.html", tp_l)
        html_c = f"""
        Estimado cliente, adjuntamos en este correo, su factura correspondiente de {self.bsobj.nombrefactura}. Gracias por su preferencia.
        """

        file_name = os.path.basename(pdf_file)
        file_name_xml = os.path.basename(xml_file)
        logging.info(f'Enviando correo a {docobj.pdv_email} con factura {docobj.doc_numero}')
        subject = f'{docobj.bs} {tipo} {docobj.doc_numero} generada'

        # Enviar via API de Mailgun (SMTP bloqueado en Digital Ocean)
        mailgun_domain = os.environ.get('MAILGUN_DOMAIN', 'm.altamachines.com')
        mailgun_api_key = os.environ.get('MAILGUN_API_KEY', '')

        with open(pdf_file, 'rb') as f_pdf, open(xml_file, 'rb') as f_xml:
            response = requests.post(
                f"https://api.mailgun.net/v3/{mailgun_domain}/messages",
                auth=("api", mailgun_api_key),
                files=[
                    ("attachment", (file_name, f_pdf, 'application/pdf')),
                    ("attachment", (file_name_xml, f_xml, 'application/xml')),
                ],
                data={
                    "from": settings.DEFAULT_FROM_EMAIL,
                    "to": [docobj.pdv_email],
                    "subject": subject,
                    "html": html_c,
                },
                timeout=30
            )

        if response.status_code != 200:
            logging.error(f'Error enviando email: {response.status_code} - {response.text}')
            raise Exception(f'Error Mailgun: {response.status_code} - {response.text}')

        docobj.enviado_cliente = True
        docobj.enviado_cliente_fecha = datetime.now()
        docobj.save()
        return {'success': f'Notificado al cliente factura {docobj.doc_numero}'}, args, kwargs


    def send_invoice_direct(self, *args: list, **kwargs: dict) -> tuple:
        q: dict = kwargs.get('qdict', {})
        force_send = q.get('force_send', False)
        userobj = kwargs.get('userobj')
        if not userobj:
            userobj = User.objects.get(username='amadmin')
        docpk = q.get('docpk')
        dbcon = q.get('dbcon')
        from_console = q.get('from_console')
        docobj = DocumentHeader.objects.get(pk=docpk)
        logging.info(f'Enviando factura {docobj.doc_numero} al cliente {docobj.pdv_email}')
        tipo = 'Factura'
        if docobj.doc_tipo in ['NC']:
            tipo = 'Nota de credito'        
        if not force_send:
            if docobj.enviado_cliente  and from_console:
                return {'error': 'El correo ya fue enviado al cliente '}

        dattrs = {'media_path': True}

        # Get Business object data based on ek_bs_ruc
        bobj = Business.objects.get(ruc=docobj.ek_bs_ruc)
        bobj_dict = model_to_dict(bobj, exclude=['contribuyenteobj', 'actividadecoobj'])
        bobj_dict['contribuyente'] = bobj.contribuyenteobj.tipo
        bobj_dict['ciudad'] = bobj.ciudadobj.nombre_ciudad
        bobj_dict['denominacion'] = bobj.actividadecoobj.nombre_actividad
        # Convert ciudadobj FK to name string
        dattrs.update(bobj_dict)

        # Add logo from configuration
        #dattrs['sifen_logo'] = fl_sifen_conf.LOGO
        logging.info(f'Generando PDF para factura {docobj.doc_numero}')
        d_pdf = self.crear_documentheader_pdf(
            userobj=userobj,
            qdict={
                'dbcon': dbcon,
                'id': docpk,
                'dattrs': to_json(dattrs)
        })
        pdf_file = f'{settings.BASE_DIR}/{d_pdf.get("pdf_file")}'
        logging.info(f'PDF generado en {pdf_file}')
        xml_file = docobj.ek_xml_file_signed.path
        tp_l = {
            "tipo": tipo,
            "razon_social": docobj.pdv_nombrefactura,
            "ruc": docobj.pdv_ruc,
            "monto": docobj.get_total_venta(),
            "link": docobj.ek_qr_link,
            "numero": '{}-{}-{}'.format(
                str(docobj.doc_establecimiento).zfill(3),
                str(docobj.doc_expedicion).zfill(3),
                str(docobj.doc_numero).zfill(7),
            )
        }

        #html_c = render_to_string("Sifen/DocumentHeaderEmailUi.html", tp_l)
        html_c = f"""
        Estimado cliente, adjuntamos en este correo, su factura correspondiente de {self.bsobj.nombrefactura}. Gracias por su preferencia.
        """
        #txt_c = strip_tags(html_c)

        dbcon = q.get('dbcon')
        with open(pdf_file, 'rb') as file:
            file_content = file.read()
        with open(xml_file, 'rb') as file:
            file_content_xml = file.read()
        # Extract the filename from the attachment_path
        file_name = os.path.basename(pdf_file)
        file_name_xml = os.path.basename(xml_file)
        logging.info(f'Enviando correo a {docobj.pdv_email} con factura {docobj.doc_numero}')
        subject = f'{docobj.bs} {tipo} {docobj.doc_numero} generada'
        email = EmailMultiAlternatives(subject=subject,
                                    body=subject,
                                    from_email=settings.DEFAULT_FROM_EMAIL,
                                    to=[docobj.pdv_email])
        email.attach_alternative(html_c, 'text/html')
        email.attach(file_name, file_content, 'application/pdf')
        email.attach(file_name_xml, file_content_xml, 'application/xml')
        email.send()
        docobj.enviado_cliente = True
        docobj.enviado_cliente_fecha = datetime.now()
        docobj.save()
        return {'success': f'Notificado al cliente factura {docobj.doc_numero}'}, args, kwargs        

    def validate_solicitud(self, 
                           tipo: str,
                           source: Union[str, None], 
                           ext_link: Union[str, int] , 
                           details: list, pagos: list,
                           cre_tipo_cod: int
                           ):
        logging.info(f'Validando solicitud de proforma {tipo} {source} {ext_link}')
        if DocumentHeader.objects.filter(
            doc_tipo=tipo,
            source=source,
            ext_link=ext_link):
            msg_e = f'La solicitud con origen {source} y id {ext_link} ya existe'
            logging.error(msg_e)
            return {'error': msg_e}
        if not details:
            msg_e = f'Debe de especificar el detalle'
            logging.error(msg_e)
            return {'error': msg_e}
        vals = set(['prod_cod', 'prod_descripcion', 'precio_unitario', 'cantidad'])
        derrs = []
        for d in details:
            kks = set(d.keys())
            for ve in  vals.difference(kks):
                derrs.append(f'Falta el campo {ve}')
        if derrs:
            msg_e = '\n'.join(derrs)
            logging.error(msg_e)
            return {'error': msg_e}
        if tipo in ['FE']:
            if not pagos and cre_tipo_cod != 2:
                msg_e = f'Debe de especificar los pagos'
                logging.error(msg_e)
                return {'error': msg_e}
            perrs = []
            for p in pagos:
                kks = p.keys()
                if int(p.get('tipo_cod')) in [ 2 ]:
                    if 'cheque_numero' not in kks:
                        perrs.append(f' Falta el campo cheque_numero')
                    if 'cheque_emisor' not in kks:
                        perrs.append(f' Falta el campo cheque_emisor')
                if int(p.get('tipo_cod')) in [  3, 4 ]:
                    if 'tarjeta_denominacion_cod' not in kks:
                        perrs.append(f' Falta el campo tarjeta_denominacion_cod')
                    if 'tarjeta_procesadora' not in kks:
                        perrs.append(f' Falta el campo tarjeta_procesadora')
                    if 'tarjeta_procesadora_ruc' not in kks:
                        perrs.append(f' Falta el campo tarjeta_procesadora_ruc')
                    if 'tarjeta_procesadora_ruc_dv' not in kks:
                        perrs.append(f' Falta el campo tarjeta_procesadora_ruc_dv')
                    if 'tarjeta_procesamiento' not in kks:
                        perrs.append(f' Falta el campo tarjeta_procesamiento')
                    if 'tarjeta_autorizacion_cod' not in kks:
                        perrs.append(f' Falta el campo tarjeta_autorizacion_cod')
                    if 'tarjeta_titular' not in kks:
                        perrs.append(f' Falta el campo tarjeta_titular')
                    if 'tarjeta_numero' not in kks:
                        perrs.append(f' Falta el campo tarjeta_numero')
            if perrs:
                return {'error': '\n'.join(perrs)}
        return {'success': 'Hecho'}

    def crear_documentheader(self, *args: list, **kwargs: dict) -> tuple:
        ios = IoS()
        eser = ekuatia_serials.Eserial()
        userobj = kwargs.get('userobj')
        q: dict = kwargs.get('qdict', {})
        dbcon = q.get('dbcon')
        uc_fields: dict = from_json(q.get('uc_fields', "{}"))
        send_invoice = uc_fields.pop('send_invoice', None)
        insert_pago = uc_fields.pop('insert_pago', None)
        not_redon = uc_fields.pop('not_redon', None)
        ti_pagos = uc_fields.pop('ti_pagos', None)
        uc_fields['doc_numero'] = 0
        uc_fields['cargado_usuario'] = userobj.first_name
        uc_fields['cargado_fecha'] = datetime.now()
        uc_fields['observacion'] = uc_fields['observacion'].strip()
        detail = uc_fields.pop('detail')
        ff = ios.form_model_fields(uc_fields, DocumentHeader._meta.fields)
        for rr in ff:
            uc_fields.pop(rr)
        rnorm, rrm, rbol = ios.format_data_for_db(DocumentHeader, uc_fields)
        for c in rrm: uc_fields.pop(c)
        for f, fv in rbol: uc_fields[f] = fv
        for f, fv in rnorm: uc_fields[f] = fv
        clobj = Clientes.objects.using('fl').get(clientecodigo=uc_fields['pdv_codigo'])
        doc_establecimiento = fl_sifen_conf.ESTABLECIMIENTO.get(clobj.sucursal.sucursal)
        if not doc_establecimiento: return {'error': 'Establecimiento incorrecto'}, args, kwargs
        doc_numero = eser.get_last_number(qdict={
            'establecimiento': doc_establecimiento,
            'tipo': uc_fields['doc_tipo']
        })
        if doc_numero.get('error'):
            return {'error': 'No existen numeros disponibles'}, args, kwargs
        
        #Datos del cliente
        tipo_cliente = clobj.tipo_cliente if clobj.tipo_cliente else 1
        
        uc_fields['pdv_tipocontribuyente'] = tipo_cliente
        uc_fields['pdv_es_contribuyente'] =  True if tipo_cliente else False
        uc_fields['pdv_type_business'] = fl_sifen_conf.K_TIPO_CON.get(tipo_cliente)
        uc_fields['pdv_nombrefantasia'] = clobj.normalize_name()
        uc_fields['pdv_nombrefactura'] = clobj.razon_social
        uc_fields['pdv_celular'] = clobj.clientecelular
        uc_fields['pdv_pais_cod'] = 'PRY'
        uc_fields['pdv_pais'] = 'Paraguay'
            
        uc_fields['doc_numero'] = None
        
        uc_fields['doc_cre_tipo'] = fl_sifen_conf.K_CRE_TIPO_COD.get(int(uc_fields['doc_cre_tipo_cod']))
        #Tipo de operacion
        uc_fields['doc_tipo_ope'] =  2
        uc_fields['doc_tipo_ope_desc'] =  u'Prestación de servicios'
        uc_fields['doc_op_pres_cod'] = 1
        uc_fields['doc_op_pres'] = u"Operación presencial"
        #Datos de pago
        uc_fields['doc_tipo_pago_cod'] = 1  # iTiPago Efectivo, Cheque, TC, TD, VALE ...etc
        uc_fields['doc_tipo_pago'] = fl_sifen_conf.K_TIPO_PAGO.get(1)
        uc_fields['doc_establecimiento'] = doc_establecimiento
        uc_fields['doc_establecimiento_ciudad'] = fl_sifen_conf.ESTABLECIMIENTO_CIUDAD.get(clobj.sucursal.sucursal)
        uc_fields['ek_timbrado'] = doc_numero.get('timbrado')
        uc_fields['bs'] = self.bsobj.name
        uc_fields['ek_bs_ruc'] = self.bsobj.ruc
        uc_fields['ek_bs_ae'] = self.bsobj.actividadecoobj.nombre_actividad
        uc_fields['ek_bs_ae_cod'] = self.bsobj.actividadecoobj.codigo_actividad
        uc_fields['ek_idcsc'] = self.timbradoobj.fcsc
        uc_fields['ek_idscsc'] = self.timbradoobj.scsc
        uc_fields['ek_timbrado_vigencia'] = self.timbradoobj.inicio
        uc_fields['ek_timbrado_vencimiento'] = self.timbradoobj.vencimiento

        uc_fields['doc_vencimiento'] = arrow.get().shift(days=30).strftime('%Y-%m-%d')
        uc_fields['doc_total'] = 0
        uc_fields['doc_iva'] = 0
        uc_fields['doc_exenta'] = 0
        uc_fields['doc_g10'] = 0
        uc_fields['doc_i10'] = 0
        uc_fields['doc_g5'] = 0
        uc_fields['doc_i5'] = 0
        uc_fields['doc_descuento'] = 0
        uc_fields['doc_per_descuento'] = 0
        uc_fields['doc_descuento_global'] = 0
        uc_fields['doc_saldo'] = 0
        uc_fields['doc_pago'] = 0
        uc_fields['doc_costo'] = 0
        uc_fields['doc_redondeo'] = 0
        uc_fields['peso'] = 0
        uc_fields['volumen'] = 0
        if uc_fields['doc_moneda'] == 'USD':
            uc_fields['doc_moneda'] = 'GS'
        if uc_fields['doc_tipo'] in ['NC', 'ND']:
            docheaderobj_rel = DocumentHeader.objects.get(pk=uc_fields['doc_relacion_cdc'])
            uc_fields['doc_loop_link'] = docheaderobj_rel.prof_number
            uc_fields['doc_relacion'] = 'Electrónico'
            uc_fields['doc_relacion_cod'] = 1 # 1 = Electronico, 2 = Impreso 3 = Constancia Eleectronica
            uc_fields['doc_relacion_cdc'] = docheaderobj_rel.ek_cdc
            uc_fields['doc_relacion_timbrado'] = docheaderobj_rel.ek_timbrado
            uc_fields['doc_relacion_establecimiento'] = docheaderobj_rel.doc_establecimiento
            uc_fields['doc_relacion_numero'] = docheaderobj_rel.doc_numero
            uc_fields['doc_relacion_expedicion'] = docheaderobj_rel.doc_expedicion
            uc_fields['doc_relacion_tipo_cod'] =  docheaderobj_rel.doc_tipo_ope
            uc_fields['doc_relacion_tipo'] = docheaderobj_rel.doc_tipo_ope_desc
            uc_fields['doc_relacion_fecha'] = docheaderobj_rel.doc_fecha
            uc_fields['doc_relacion_monto'] = docheaderobj_rel.doc_total_redondeo
            uc_fields['doc_op'] = 'NC'
            uc_fields['doc_estado'] = 'DESCUENTO'
            uc_fields['doc_motivo'] = 'Bonificación'
            uc_fields['doc_tipo_ope'] = 2
            uc_fields['doc_tipo_ope_desc'] =  u'Prestación de servicios'
            uc_fields['doc_tipo_imp'] = 1
            uc_fields['doc_op_pres_cod'] = 1
            uc_fields['doc_op_pres'] = u"Operación presencial"
            uc_fields['doc_cre_tipo_cod'] = 2
            uc_fields['doc_cre_tipo'] = fl_sifen_conf.K_CRE_TIPO_COD.get(2)
            uc_fields['doc_tipo_pago_cod'] = 1
        if uc_fields['doc_tipo'] == "AF":
            uc_fields['doc_relacion'] = u'Constancia Electr\xf3nica'
            uc_fields['doc_relacion_cod'] = 3
            uc_fields['doc_relacion_tipo_cod'] =  1
            uc_fields['doc_relacion_tipo'] = u'Constancia de no ser contribuyente'
            uc_fields['af_vendedor'] = fl_sifen_conf.K_VENDEDOR_COD.get(int(uc_fields['af_vendedor_cod']))
            uc_fields['af_tdoc'] = fl_sifen_conf.K_TDOC_COD.get(int(uc_fields['af_tdoc_cod']))
            uc_fields['af_nro_casa'] = 0
            uc_fields['af_dpto_nombre'] = Departamentos.objects.get(codigo_departamento=uc_fields['af_dpto_cod']).nombre_departamento
            uc_fields['af_distrito_nombre'] = Distrito.objects.get(codigo_distrito=uc_fields['af_distrito_cod']).nombre_distrito
            uc_fields['af_ciudad_nombre'] = Ciudades.objects.get(codigo_ciudad=uc_fields['af_ciudad_cod']).nombre_ciudad
            uc_fields['pdv_direccion_entrega'] = self.bsobj.direccion
        if not uc_fields.get('doc_expedicion'):
            uc_fields['doc_expedicion'] = 1
        if ti_pagos:
            uc_fields['forma_pago_id'] = ti_pagos.get('forma_pago_id')
            uc_fields['forma_pago'] = ti_pagos.get('forma_pago')
        docobj = DocumentHeader.objects.using(dbcon).create(**uc_fields)
        if docobj.doc_cre_tipo_cod == 2:
            docobj.doc_cre_cond = 1
            docobj.doc_cre_plazo='30 dias'
            docobj.save()
        files: dict = kwargs.get('files')
        for name, fobj in files.items():
            fname = f'op_{fobj.name}'
            dfobj = File(fobj, name=fname)
            setattr(docobj, name, dfobj)
        docobj.save()
        times_non_prod = 0
        doc_redondeo = 0
        for d in detail:
            prod_autocreado = False
            if docobj.doc_tipo == "AF":
                PAF = namedtuple('PAF', ['prod_cod','precio', 'moneda', 'g5', 'g10', 'exenta', 'volumen', 'peso',
                                         'medidaobj','porcentaje_iva' ])
                MOBJ = namedtuple('MOBJ', ['medida_cod', 'medida'])
                IOBJ = namedtuple('IOBJ', ['porcentaje'])
                medidaobj = MOBJ(medida_cod=77, medida='UNI')
                porcentajeobj = IOBJ(porcentaje=10)
                prod_autocreado = True
                prodobj = PAF(
                    prod_cod=999,
                    precio=d.get('precio_unitario'), 
                    moneda='GS', 
                    g5=0, 
                    g10=100, 
                    exenta=0, 
                    volumen=0, 
                    peso=0,
                    medidaobj=medidaobj,
                    porcentaje_iva=porcentajeobj,
                )
            else:
                if int(d.get('prod_cod')) == 1850:
                    prod_autocreado = True
                    times_non_prod += 50
                    PAF = namedtuple('PAF', ['prod_cod','precio', 'moneda', 'g5', 'g10', 'exenta', 'volumen', 'peso',
                                         'medidaobj','porcentaje_iva' ])
                    MOBJ = namedtuple('MOBJ', ['medida_cod', 'medida'])
                    IOBJ = namedtuple('IOBJ', ['porcentaje'])
                    medidaobj = MOBJ(medida_cod=77, medida='UNI')
                    porcentajeobj = IOBJ(porcentaje=10)
                    prodobj = PAF(
                        prod_cod=1850+times_non_prod,
                        precio=d.get('precio_unitario'), 
                        moneda='GS', 
                        g5=0, 
                        g10=100, 
                        exenta=0, 
                        volumen=0, 
                        peso=0,
                        medidaobj=medidaobj,
                        porcentaje_iva=porcentajeobj,
                    )
                else:
                    prodobj = Producto.objects.get(prod_cod=d.get('prod_cod'))
                    #Articulos migrados de la base de datos antigua no tienen precio
                    if prodobj.precio <= 0 or d.get('not_recal') == True:
                        prodobj.precio = Decimal(d.get('precio_unitario'))
                        if d.get('change_currency') == True:
                            prodobj.moneda = docobj.doc_moneda
            per_tipo_iva = 100
            precio = prodobj.precio
            precio_source = prodobj.precio
            docobj.tasa_cambio = Decimal(docobj.tasa_cambio)
            if docobj.doc_tipo == 'NC':
                precio = Decimal(d.get('precio_unitario'))
            else:
                if ((docobj.doc_moneda == 'GS') and (prodobj.moneda == 'USD')):
                    precio = int(precio * docobj.tasa_cambio)
                if ((docobj.doc_moneda == 'USD') and (prodobj.moneda == 'GS')):
                    precio = int(precio / docobj.tasa_cambio)
                if ((docobj.doc_moneda == 'GS') and (prodobj.moneda == 'GS')):
                    precio = int(precio)
            pcalc = f_calcs.calculate_price(prodobj, 
                    precio,
                    d.get('cantidad')
            )
            pcalc_source = pcalc.copy()
            pcalc_source['afecto'] = pcalc_source.get('gravada_5', 0)+pcalc_source.get('gravada_10', 0)
            pd_total = float(pcalc.get('exenta', 0)+pcalc.get('gravada_10', 0)+pcalc.get('gravada_5', 0))
            v_redondeo = pd_total
            #print(v_redondeo)
            if docobj.doc_moneda == 'GS' and not_redon == None:
                v_redondeo = f_calcs.round_to_nearest_50(pd_total)
                doc_redondeo += (pd_total - v_redondeo)*-1
                #print(v_redondeo, doc_redondeo, 'verga')
            #print(v_redondeo, pd_total, 'verga', doc_redondeo)
            if v_redondeo != pd_total:
                precio = v_redondeo/d.get('cantidad')
                pcalc = f_calcs.calculate_price(prodobj, 
                    precio,
                    d.get('cantidad'),
                )
            if pcalc.get('exenta') and pcalc.get('afecto'):
                #total = pcalc.get('exenta')+pcalc.get('afecto')
                #per_tipo_iva = (float(prodobj.exenta)*100)/float(total)
                per_tipo_iva = prodobj.g5+prodobj.g10
            #print(pcalc, per_tipo_iva)
            pcalc['afecto'] = pcalc.get('gravada_5', 0)+pcalc.get('gravada_10', 0)
            docobj.documentdetail_set.create(
                prod_autocreado = prod_autocreado,
                prod_cod = prodobj.prod_cod,
                prod_descripcion = d.get('prod_descripcion'),
                prod_unidad_medida = prodobj.medidaobj.medida_cod,
                prod_unidad_medida_desc = prodobj.medidaobj.medida,
                prod_pais_origen = d.get('prod_pais_origen', 'PRY'),
                prod_pais_origen_desc = d.get('prod_pais_origen_desc', 'Paraguay'),
                prod_lote = d.get('prod_lote', None),
                prod_vencimiento = d.get('prod_vencimiento', None),
                porcentaje_iva = prodobj.porcentaje_iva.porcentaje,
                precio_unitario_source = precio_source,
                precio_unitario = pcalc.get('precio_unitario'),
                precio_unitario_siniva = pcalc.get('precio_unitario_siniva'),
                cantidad = d.get('cantidad'),
                cantidad_devolucion = 0,
                #CALCULAR SEGUN PRODUCTO
                exenta = pcalc.get('exenta'), 
                iva_5 = pcalc.get('iva_5'), 
                gravada_5 = pcalc.get('gravada_5'), 
                base_gravada_5 = pcalc.get('base_gravada_5'),
                iva_10 = pcalc.get('iva_10'), 
                gravada_10 = pcalc.get('gravada_10'), 
                base_gravada_10 = pcalc.get('base_gravada_10'),
                afecto = (pcalc.get('gravada_5', 0)+pcalc.get('gravada_10', 0)),
                per_tipo_iva = per_tipo_iva,
                bonifica = False,
                descuento = 0,
                per_descuento = 0,
                volumen = prodobj.volumen,
                peso = prodobj.peso,
                observacion = None,
                cargado_usuario=userobj.first_name,
                cargado_fecha=datetime.now(),
                pcalc_source=from_json(to_json(pcalc_source)),
                pcalc_result=from_json(to_json(pcalc))
            )
        # update totals
        totals = docobj.documentdetail_set.aggregate(
                            exenta=Sum('exenta'),
                            iva_5=Sum('iva_5'),
                            gravada_5=Sum('gravada_5'),
                            iva_10=Sum('iva_10'),
                            gravada_10=Sum('gravada_10'),
                            afecto=Sum('afecto')
                        )
        doc_total = totals.get('afecto') + totals.get('exenta')

        docobj.doc_total = doc_total
        docobj.doc_iva = totals.get('iva_5') + totals.get('iva_10')
        docobj.doc_exenta = totals.get('exenta')
        docobj.doc_g10 = totals.get('gravada_10')
        docobj.doc_i10 = totals.get('iva_10')
        docobj.doc_g5 = totals.get('gravada_5')
        docobj.doc_i5 = totals.get('iva_5')
        docobj.doc_redondeo = doc_redondeo
        if docobj.doc_cre_tipo_cod == 2:
            docobj.doc_saldo = docobj.get_total_venta_gs()
        docobj.save()
        total_venta_gs = docobj.get_total_venta_gs()
        docobj.doc_total_redondeo = docobj.get_total_operacion_redondeo()

        if docobj.doc_tipo == 'FE':
            docobj.doc_relacion_saldo = total_venta_gs
        if docobj.doc_tipo == 'NC':
            if total_venta_gs > docheaderobj_rel.doc_relacion_saldo:
                docobj.delete()
                return {'error': 'El valor de la NC {:,.2f} es superior al saldo de la factura {:,.2f}'.format(total_venta_gs, docheaderobj_rel.doc_relacion_saldo) }, args, kwargs
            docheaderobj_rel.doc_relacion_saldo = docheaderobj_rel.doc_relacion_saldo - total_venta_gs
            docheaderobj_rel.save()
        
        docobj.save()
        #print('FUCK TREE')
        rsp = self.set_number(
            [docobj.prof_number],
            1,
            int(doc_establecimiento),
            sign_document=True,
            doc_tipo=docobj.doc_tipo
        )
        if not rsp.get('success'):
            return {'error': 'No fue posible firmar el documento'}, args, kwargs
        if send_invoice == True:
            send_task('Sifen.tasks.send_invoice', 
                kwargs={
                    'username': userobj.username,
                    'qdict': {
                        'dbcon': dbcon,
                        'docpk': docobj.id,
                        'dattrs': to_json({
                            'full': 1,
                        })
                }
            })
        return {'success': 'Documento creado', 'record_id': docobj.id}, args, kwargs

    def generando_documentheader(self, *args: list, **kwargs: dict) -> dict:
        userobj = kwargs.get('userobj')
        q: dict = kwargs.get('qdict', {})
        dbcon = q.get('dbcon')
        id = q.get('id')

        # Get document to access ek_bs_ruc
        docobj = DocumentHeader.objects.using(dbcon).get(pk=id)

        # Initialize dattrs
        dattrs = {'media_path': True}

        # Get Business object data based on ek_bs_ruc
        bobj = Business.objects.get(ruc=docobj.ek_bs_ruc)
        bobj_dict = model_to_dict(bobj, exclude=['contribuyenteobj', 'actividadecoobj'])
        bobj_dict['contribuyente'] = bobj.contribuyenteobj.tipo
        bobj_dict['ciudad'] = bobj.ciudadobj.nombre_ciudad
        bobj_dict['denominacion'] = bobj.actividadecoobj.nombre_actividad
        # Convert ciudadobj FK to name string
        dattrs.update(bobj_dict)

        # Add logo from configuration
        #dattrs['sifen_logo'] = fl_sifen_conf.LOGO

        #print(fl_sifen_conf.LOGO)

        return self.crear_documentheader_pdf(
            username=userobj.username,
            userobj=userobj,
            qdict={
                'dbcon': dbcon,
                'id': id,
                'dattrs': to_json(dattrs)
        })

        # self.crear_documentheader_pdf(
        #     username=userobj.username,
        #     userobj=userobj,
        #     qdict={
        #         'dbcon': dbcon,
        #         'id': id,
        #         'dattrs': to_json({
        #             'title_header': 'Orden de Pedido',
        #             'base_cond': c
        #         })
        # })

    @tracktask
    def crear_documentheader_pdf(self, *args: list, **kwargs: dict) -> dict:
        q: dict = kwargs.get('qdict')
        dbcon = q.get('dbcon')
        id =  q.get('id')
        dattrs = q.get('dattrs')
        docobj = DocumentHeader.objects.using(dbcon).get(id=id)
        mq = QueryDict(mutable=True)
        fv = {''
            'tmpl': 'Sifen/DocumentHeaderRptUi.html',
            'model_app_name':'Sifen',
            'model_name':'DocumentHeader',
            'pk':id,
            'dattrs': dattrs,
            'dbcon':'default',
            'surround': 'BaseAmReportUi.html',
            'g_pdf': 0,
            'g_pdf_kit': 1,
            'page-size': 'A5',
            'orientation': 'Portrait',
            'page-height': '250.000212',
            'page-width': '210.000088',
            'margin-top': '0',
            'margin-right': '0',
            'margin-bottom': '0',
            'margin-left': '0',
            'no-outline': None
            #'orientation': 'Landscape',
        }
        mq.update(fv)
        mrq = HttpRequest()
        mrq.user = kwargs.get('userobj')
        mrq.method = 'GET'
        mrq.GET = mq
        io_rpt = IoRpt()
        rp = io_rpt.rpt_view(mrq)
        pdf_file = rp.get('pdf_file')
        full = False
        if dattrs:
            full = from_json(dattrs).get('full')
        #NOTE: Se decide generar el PDF siempre bajo demanda
        with open(pdf_file, 'rb') as f:
           docobj.ek_pdf_file = File(f, name=os.path.basename(pdf_file))
           docobj.save()
        print(rp.get('html_file'))

        # Convert full path to media URL
        # pdf_file is now in format: /path/to/project/media/rpt/12345.pdf
        # We need to return: /media/rpt/12345.pdf
        pdf_filename = os.path.basename(pdf_file)
        pdf_media_url = f'{settings.MEDIA_URL}rpt/{pdf_filename}'

        return {'success': f'Pdf del documento {docobj.id} creado con exito',
                'pdf_file': pdf_media_url,
                'ek_pdf_file': pdf_media_url,
                'html_file': rp.get('html_file')
            }

    def crear_proforma(self,
            userobj,
            clientecodigo: int = 0,
            expedicion: int = 1,
            source: Union[str, None] = None, # Modulo de donde viene la peticion
            ext_link: Union[str,int] = 0, # Identificadoor del modelo de donde se genera la solicitud
            doc_moneda: Literal['GS', 'USD', 'EURO'] = 'GS',
            doc_fecha: str = arrow.get().date().strftime('%Y-%m-%d'),
            doc_tipo: Literal['FE', 'NC', 'ND', 'AF'] = 'FE', #Describe el tipo de documento
            doc_op: Literal['VTA', 'RETIRO', 'NC'] = 'VTA', #Describe la operacion
            doc_estado: str = 'ND', #El estado de la operacion
            doc_vencimiento: str = arrow.get().shift(days=30).strftime('%Y-%m-%d'), #Cuando vence la factura
            doc_total: float = 0, #Valor total de la operacion
            doc_iva: float = 0, #Impuesta sobre el valor total
            doc_exenta: float = 0, #Valor exenta
            doc_g10: float = 0, #Gravada 10
            doc_i10: float = 0, #Iva 10
            doc_g5: float = 0, #Gravada 5
            doc_i5: float = 0, #Iva 5
            doc_descuento: float = 0, #Valor del descuento sobre el total
            doc_per_descuento: float = 0, #Porcentaje del descuento sobre el total
            doc_tipo_ope: int = 2, # Referencia al tipo de operacion del Sifen
            doc_tipo_imp: int = 1,
            op_pres_cod: int = 1,
            cre_tipo_cod: int = 1, # 1 = Contado 2 = Crédito
            doc_tipo_pago_cod: int = 1, #Efectivo
            doc_cre_cond: Union[int, None] = None, #
            doc_cre_plazo: Union[str, None] = None,
            doc_cre_cuota: Union[str, None] = None,
            doc_cre_entrega_inicial: Union[int, None] = None,
            pdv_pais_cod: str = 'PRY',
            pdv_pais: str = 'Paraguay',
            pdv_direccion_entrega: Union[str, None] = None,
            pdv_dir_calle_sec: Union[str, None] = None,
            pdv_direccion_comple: Union[str, None] = None,
            pdv_numero_casa: int = 0,
            pdv_numero_casa_entrega: int = 0,
            pdv_dpto_cod: int = 0,
            pdv_dpto_nombre: Union[str, None] = None,
            pdv_distrito_cod: int = 0,
            pdv_distrito_nombre: Union[str, None] = None,
            pdv_ciudad_cod: int = 0,
            pdv_ciudad_nombre: Union[str, None] = None,
            doc_loop_link: Union[str, uuid.UUID, None] = None,
            details: list = [],
            pagos: list = [],
            doc_redondeo: float = 0,
            doc_voucher: float = 0,
            not_redon = None
        ):
        logging.info('Crear proforma')
        cg = Cotizacion.objects.using('fl').order_by('id_cotizacion').last()
        if cg:
            tasa_cambio = cg.dolar_venta
        logging.info(f'Cotizacion del dia {tasa_cambio}')
        eser = ekuatia_serials.Eserial()
        gdata = mng_gmdata.Gdata()
        rsp = self.validate_solicitud(
                    doc_tipo,
                    source, 
                    ext_link, 
                    details, 
                    pagos,
                    cre_tipo_cod
        )
        if rsp.get('error'):
            return rsp
        
        pdv_innominado = False
        tipo_cliente = 1
        pdv_ruc = 'ND'
        pdv_ruc_dv = -1
        pdv_nombrefantasia = 'INNOMINADO'
        pdv_nombrefactura = 'INNOMINADO'
        pdv_telefono = None
        pdv_celular = None
        pdv_email = None
        
        if clientecodigo:
            clobj = Clientes.objects.using('fl').get(clientecodigo=clientecodigo)
            tipo_cliente = clobj.tipo_cliente
            pdv_ruc = clobj.get_ruc()
            pdv_ruc_dv = gdata.calculate_dv(pdv_ruc)
            pdv_nombrefantasia = clobj.normalize_name()
            pdv_nombrefactura = clobj.razon_social
            pdv_telefono = clobj.clientetelefono
            pdv_celular = clobj.clientecelular
            pdv_email = clobj.clientemail
            doc_establecimiento = fl_sifen_conf.ESTABLECIMIENTO.get(clobj.sucursal.sucursal)
            if not doc_establecimiento: return {'error': 'Establecimiento incorrecto'}
        else:
            doc_establecimiento = fl_sifen_conf.ESTABLECIMIENTO.get(1)
        doc_numero = eser.get_last_number(qdict={
            'establecimiento': doc_establecimiento,
            'tipo': doc_tipo
        })
        if doc_numero.get('error'):
            return {'error': 'No existen numeros disponibles'}, args, kwargs
        doc_relacion=None
        doc_relacion_cod=0
        doc_relacion_cdc=None
        doc_relacion_timbrado=None
        doc_relacion_establecimiento=0
        doc_relacion_numero=0
        doc_relacion_expedicion=0
        doc_relacion_tipo=None
        doc_relacion_tipo_cod=0
        doc_relacion_fecha=None
        doc_relacion_monto=0
        if doc_op:
            doc_relacion_tipo_cod = 1
            doc_relacion_tipo = 'Constancia de no ser contribuyente'
        if doc_op in ['NC', 'ND']:
            docheaderobj_rel = DocumentHeader.objects.get(prof_number=doc_loop_link)
            doc_relacion = 'Electrónico'
            doc_relacion_cod = 1 # 1 = Electronico, 2 = Impreso 3 = Constancia Eleectronica
            doc_relacion_cdc = docheaderobj_rel.ek_cdc
            doc_relacion_timbrado = docheaderobj_rel.ek_timbrado
            doc_relacion_establecimiento = docheaderobj_rel.doc_establecimiento
            doc_relacion_numero = docheaderobj_rel.doc_numero
            doc_relacion_expedicion = docheaderobj_rel.doc_expedicion
            doc_relacion_tipo_cod =  docheaderobj_rel.doc_tipo_ope
            doc_relacion_tipo = docheaderobj_rel.doc_tipo_ope_desc
            doc_relacion_fecha = docheaderobj_rel.doc_fecha
            doc_relacion_monto = docheaderobj_rel.doc_total_redondeo
        docheaderobj = DocumentHeader.objects.create(
            bs=self.bsobj.name,
            source=source,
            ext_link=ext_link,
            doc_moneda=doc_moneda,
            doc_fecha=doc_fecha,
            doc_tipo=doc_tipo,
            doc_op=doc_op,
            doc_estado=doc_estado,
            doc_vencimiento=doc_vencimiento,
            doc_total=doc_total,
            doc_iva=doc_iva,
            doc_exenta=doc_exenta,
            doc_g10 = doc_g10,
            doc_i10 = doc_i10,
            doc_g5 = doc_g5,
            doc_i5 = doc_i5,
            doc_descuento = doc_descuento,
            doc_per_descuento = doc_per_descuento,
            doc_descuento_global = 0,
            doc_saldo = 0,
            doc_pago = 0,
            doc_costo = 0,
            doc_redondeo = doc_redondeo,
            doc_establecimiento = fl_sifen_conf.ESTABLECIMIENTO.get(clobj.sucursal.sucursal),
            doc_expedicion = expedicion,
            doc_tipo_ope = doc_tipo_ope,
            doc_tipo_ope_desc = fl_sifen_conf.K_TIPO_OPE.get(doc_tipo_ope),
            doc_tipo_imp = doc_tipo_imp,
            doc_tipo_imp_desc = fl_sifen_conf.K_TIPO_IMPUESTO.get(doc_tipo_imp),
            doc_op_pres_cod = op_pres_cod,
            doc_op_pres = fl_sifen_conf.K_OP_RES.get(op_pres_cod),
            doc_cre_tipo_cod = cre_tipo_cod,
            doc_cre_tipo = fl_sifen_conf.K_CRE_TIPO_COD.get(cre_tipo_cod),
            doc_tipo_pago_cod = doc_tipo_pago_cod, 
            doc_tipo_pago = fl_sifen_conf.K_TIPO_PAGO.get(doc_tipo_pago_cod),
            doc_cre_cond=doc_cre_cond,
            doc_cre_cond_desc=fl_sifen_conf.K_CRE_COND.get(doc_cre_cond) if doc_cre_cond else None,
            doc_cre_plazo=doc_cre_plazo,
            doc_cre_cuota=doc_cre_cuota,
            doc_cre_entrega_inicial=doc_cre_entrega_inicial,
            doc_entregado=None,
            doc_entregado_usuario=None,
            doc_loop_link=doc_loop_link,
            doc_relacion=doc_relacion,
            doc_relacion_cod=doc_relacion_cod,
            doc_relacion_cdc=doc_relacion_cdc,
            doc_relacion_timbrado=doc_relacion_timbrado,
            doc_relacion_establecimiento=doc_relacion_establecimiento,
            doc_relacion_numero=doc_relacion_numero,
            doc_relacion_expedicion=doc_relacion_expedicion,
            doc_relacion_tipo=doc_relacion_tipo,
            doc_relacion_tipo_cod=doc_relacion_tipo_cod,
            doc_relacion_fecha=doc_relacion_fecha,
            doc_relacion_monto=doc_relacion_monto,
            pdv_innominado=pdv_innominado,
            pdv_pais_cod=pdv_pais_cod,
            pdv_pais=pdv_pais,
            pdv_tipocontribuyente=tipo_cliente,
            pdv_es_contribuyente= True if tipo_cliente else False,
            pdv_type_business=fl_sifen_conf.K_TIPO_CON.get(tipo_cliente),
            pdv_codigo=clientecodigo if clientecodigo else 0,
            pdv_ruc=pdv_ruc,
            pdv_ruc_dv=pdv_ruc_dv,
            pdv_nombrefantasia=pdv_nombrefantasia,
            pdv_nombrefactura=pdv_nombrefactura,
            pdv_direccion_entrega=pdv_direccion_entrega,
            pdv_dir_calle_sec=pdv_dir_calle_sec,
            pdv_direccion_comple=pdv_direccion_comple,
            pdv_numero_casa=pdv_numero_casa,
            pdv_numero_casa_entrega=pdv_numero_casa_entrega,
            pdv_dpto_cod=pdv_dpto_cod,
            pdv_dpto_nombre=pdv_dpto_nombre,
            pdv_distrito_cod=pdv_distrito_cod,
            pdv_distrito_nombre=pdv_distrito_nombre,
            pdv_ciudad_cod=pdv_ciudad_cod,
            pdv_ciudad_nombre=pdv_ciudad_nombre,
            pdv_telefono=pdv_telefono,
            pdv_celular=pdv_celular,
            pdv_email=pdv_email,
            ek_serie=self.timbradoobj.serie,
            ek_timbrado=self.timbradoobj.timbrado,
            ek_bs_ruc=self.bsobj.ruc,
            ek_bs_ae=self.bsobj.actividadecoobj.nombre_actividad,
            ek_bs_ae_cod=self.bsobj.actividadecoobj.codigo_actividad,
            ek_idcsc=self.timbradoobj.fcsc,
            ek_idscsc=self.timbradoobj.scsc,
            ek_timbrado_vigencia=self.timbradoobj.inicio,
            ek_timbrado_vencimiento=self.timbradoobj.vencimiento,
            ek_cod_seg=0,
            ek_cdc=None,
            ek_cdc_dv=0,
            ek_qr_link=None,
            ek_qr_img=None,
            ek_transacion=None,
            ek_estado=None,
            ek_date=None,
            ek_xml_ekua=False,
            ek_pdf_file=None,
            ek_xml_file=None,
            ek_xml_file_signed=None,
            impx_tdoc_cod=0,
            impx_tdoc_nam=None,
            impx_doc_num=None,
            impx_nombre=None,
            impx_cargo=None,
            tasa_cambio=tasa_cambio,
            peso=0,
            volumen=0,
            anulado_usuario=None,
            anulado_fecha=None,
            actualizado_usuario=None,
            actualizado_fecha=None
        )

        for d in details:
            pobj = Producto.objects.get(prod_cod=d.get('prod_cod'))
            #Articulo codigo 150 hace referencia al voucher
            if doc_op in ['NC', 'ND'] and pobj.prod_cod == 150:
                precio_unitario = float(d.get('precio_unitario'))
                doc_exenta = float(docheaderobj_rel.doc_exenta)
                doc_g10 = float(docheaderobj_rel.doc_g10)
                doc_g5 = float(docheaderobj_rel.doc_g5)
                total_rel = (doc_exenta + doc_g10 + doc_g5)
                per_exenta = (doc_exenta*100)/precio_unitario
                per_exenta = (doc_exenta * 100) / precio_unitario
                per_g10 = (doc_g10 * 100) / precio_unitario
                per_g5 = (doc_g5 * 100) / precio_unitario
                pcalc = {
                    'exenta': 0,
                    'gravada_10': 0,
                    'gravada_5': 0,
                    'iva_10': 0,
                    'iva_5': 0,
                    'precio_unitario_siniva': 0
                }
                if per_exenta == 0 and per_g5 == 0 and per_g10 > 0:
                    g10 = doc_g10 if precio_unitario > doc_g10 else precio_unitario
                    pcalc['afecto'] = g10
                    pcalc['gravada_10'] = g10
                    pcalc['iva_10'] = (g10 * 0.1)
                    pcalc['precio_unitario_siniva'] = precio_unitario - pcalc['iva_10']
                elif per_exenta == 0 and per_g5 > 0 and per_g10 == 0:
                    g5 = doc_g5 if precio_unitario > doc_g5 else precio_unitario
                    pcalc['afecto'] = g5
                    pcalc['gravada_5'] = g5
                    pcalc['iva_5'] = (g5 * 0.05)
                    pcalc['precio_unitario_siniva'] = precio_unitario - pcalc['iva_5']
                elif per_exenta == 0 and per_g5 > 0 and per_g10 > 0:
                    g5 = doc_g5 if precio_unitario > doc_g5 else precio_unitario
                    g10 = doc_g10 if precio_unitario > doc_g10 else precio_unitario
                    pcalc['afecto'] = g5 + g10
                    pcalc['gravada_5'] = g5
                    pcalc['gravada_10'] = g10
                    pcalc['iva_5'] = (g5 * 0.05)
                    pcalc['iva_10'] = (g10 * 0.1)
                    pcalc['precio_unitario_siniva'] = precio_unitario - (pcalc['iva_5'] - pcalc['iva_10'])
                elif per_exenta > 0 and per_g5 == 0 and per_g10 > 0:
                    if precio_unitario > total_rel:
                        pcalc['afecto'] = doc_g10
                        pcalc['gravada_10'] = doc_g10
                        pcalc['iva_10'] = (doc_g10 * 0.1)
                        pcalc['exenta'] = doc_exenta
                        pcalc['precio_unitario'] = total_rel
                        pcalc['precio_unitario_siniva'] = total_rel
                    else:
                        eper = (doc_exenta*100)/total_rel
                        gper = (doc_g10*100)/total_rel
                        g10 = (precio_unitario*gper)/100
                        pcalc['gravada_10'] = g10
                        pcalc['iva_10'] = (g10 * 0.1)
                        exe = (precio_unitario*eper)/100
                        pcalc['exenta'] = exe
                        pcalc['precio_unitario'] = precio_unitario
                        pcalc['precio_unitario_siniva'] = precio_unitario - pcalc['iva_10']
                        pcalc['afecto'] = g10
                else:
                    exe = doc_exenta if precio_unitario > doc_exenta else precio_unitario
                    pcalc['afecto'] = 0
                    pcalc['exenta'] = exe
                    pcalc['precio_unitario_siniva'] = exe
            else:
                #if prodobj.precio <= 0 
                if d.get('not_recal') == True:
                    not_redon = True
                    pobj.precio = Decimal(d.get('precio_unitario'))
                    d['cantidad'] =  1
                pcalc = f_calcs.calculate_price(pobj, 
                            d.get('precio_unitario'),
                            d.get('cantidad')
                )
            per_tipo_iva = 100
            if pcalc.get('exenta') and pcalc.get('afecto'):
                #total = pcalc.get('exenta')+pcalc.get('afecto')
                #per_tipo_iva = (float(pobj.exenta)*100)/float(total)
                per_tipo_iva = pobj.g5 + pobj.g10

            docheaderobj.documentdetail_set.create(
                prod_cod = pobj.prod_cod,
                prod_descripcion = d.get('prod_descripcion'),
                prod_unidad_medida = pobj.medidaobj.medida_cod,
                prod_unidad_medida_desc = pobj.medidaobj.medida,
                prod_pais_origen = d.get('prod_pais_origen', 'PRY'),
                prod_pais_origen_desc = d.get('prod_pais_origen_desc', 'Paraguay'),
                prod_lote = None,
                prod_vencimiento = None,
                porcentaje_iva = d.get('porcentaje_iva', 0),
                precio_unitario = d.get('precio_unitario'),
                precio_unitario_siniva = pcalc.get('precio_unitario_siniva'),
                precio_unitario_source = d.get('precio_unitario'),
                cantidad = d.get('cantidad'),
                cantidad_devolucion = d.get('cantidad_devolucion', 0),
                #CALCULAR SEGUN PRODUCTO
                exenta = pcalc.get('exenta'), 
                iva_5 = pcalc.get('iva_5'), 
                gravada_5 = pcalc.get('gravada_5'), 
                iva_10 = pcalc.get('iva_10'), 
                gravada_10 = pcalc.get('gravada_10'), 
                afecto = (pcalc.get('gravada_5')+pcalc.get('gravada_10')),
                per_tipo_iva = per_tipo_iva,
                bonifica = d.get('bonifica'),
                descuento = d.get('descuento'),
                per_descuento = d.get('per_descuento'),
                volumen = d.get('volumen', 0),
                peso = d.get('peso', 0),
                observacion = d.get('observacion', ''),
            )
        for p in pagos:
            ptmp = {
                'source': source,
                'ext_link': ext_link,
                'tipo_cod': p.get('tipo_cod'),
                'tipo': fl_sifen_conf.K_TIPO_PAGO.get(p.get('tipo_cod')),
            }
            if p.get('tarjeta_denominacion_cod'):
                ptmp['tarjeta_denominacion'] = fl_sifen_conf.K_PROCESADORA.get(p.get('tarjeta_denominacion_cod'))
            ptmp.update(p)
            docheaderobj.documentopagos_set.create(
                **ptmp
            )
        totals = docheaderobj.documentdetail_set.aggregate(
                            exenta=Sum('exenta'),
                            iva_5=Sum('iva_5'),
                            gravada_5=Sum('gravada_5'),
                            iva_10=Sum('iva_10'),
                            gravada_10=Sum('gravada_10'),
                            afecto=Sum('afecto')
                        )
        doc_total = totals.get('afecto') + totals.get('exenta')
        #doc_redondeo = 0
        if  not_redon == True:
            doc_redondeo = doc_redondeo
        if docheaderobj.doc_moneda == 'GS' and not doc_redondeo and doc_tipo in ['FE', 'ND']:
            doc_redondeo = doc_total - f_calcs.round_to_nearest_50(doc_total)
        docheaderobj.doc_total = doc_total
        docheaderobj.doc_iva = totals.get('iva_5') + totals.get('iva_10')
        docheaderobj.doc_exenta = totals.get('exenta')
        docheaderobj.doc_g10 = totals.get('gravada_10')
        docheaderobj.doc_i10 = totals.get('iva_10')
        docheaderobj.doc_g5 = totals.get('gravada_5')
        docheaderobj.doc_i5 = totals.get('iva_5')
        docheaderobj.doc_redondeo = doc_redondeo
        if docheaderobj.doc_cre_tipo_cod == 2:
            docheaderobj.doc_saldo = doc_total
        docheaderobj.doc_total_redondeo = docheaderobj.get_total_operacion_redondeo()
        if docheaderobj.doc_tipo == 'FE':
            docheaderobj.doc_relacion_saldo = docheaderobj.doc_total_redondeo
        docheaderobj.save()
        rsp = self.set_number(
            [docheaderobj.prof_number],
            1,
            int(doc_establecimiento),
            sign_document=True,
            doc_tipo=docheaderobj.doc_tipo
        )
        docheaderobj = DocumentHeader.objects.get(pk=docheaderobj.pk)
        if not rsp.get('success'):
            return {'error': 'No fue posible firmar el documento'}
        msg = f'Documento {docheaderobj.doc_tipo} {docheaderobj.doc_numero} creado con exitos'
        if doc_voucher:
            # crear el documento con el CDC asociado.
            #Traemos el articulo que representa al voucher.
            artobj = Producto.objects.get(prod_cod=150)
            rsp_asoc = self.crear_proforma(
                userobj,
                clientecodigo=clobj.clientecodigo,
                expedicion=expedicion,  # La caja donde se imprimi
                source=source,
                ext_link=ext_link,
                doc_moneda=doc_moneda,
                doc_fecha=doc_fecha,
                doc_tipo="NC",
                doc_op="NC",
                doc_estado="DESCUENTO",
                doc_vencimiento=arrow.get(doc_fecha).shift(days=16).strftime("%Y-%m-%d"),
                doc_total=doc_voucher,
                doc_tipo_ope=2,  # iTipTra 1 = Venta de mercaderia
                doc_tipo_imp=1,  # iTImp Tipo de impuesto 1 = IVA
                op_pres_cod=1,  # iIndPres como y donde se vendio
                cre_tipo_cod=1,  # iCondOpe condicion de la operacion Contado 2 = Credito
                doc_tipo_pago_cod=1,  # iTiPago Efectivo, Cheque, TC, TD, VALE ...etc
                # SOLO EN OPERACIONES DE CREDITO E640-E649
                # doc_cre_cond=1 #iCondCred 1= Plazo 2 = Cuota
                # doc_cre_plazo='30 dias' # dPlazoCre
                # doc_cre_cuota='12' #dCuotas
                # doc_cre_entrega_inicial=1200000 #dMonEnt
                pdv_pais_cod="PRY",
                pdv_pais="Paraguay",
                # No es obligatorio, pero si hay enviar, segun los modelos en la APP Sifen
                pdv_direccion_entrega=None,
                pdv_dir_calle_sec=None,
                pdv_direccion_comple=None,
                pdv_numero_casa=0,
                pdv_numero_casa_entrega=0,
                pdv_dpto_cod=0,
                pdv_dpto_nombre=None,
                pdv_distrito_cod=0,
                pdv_distrito_nombre=None,
                pdv_ciudad_cod=0,
                pdv_ciudad_nombre=None,
                # Solo en caso de operaciones que requieren un documento asociado
                doc_loop_link=docheaderobj.prof_number,
                details=[
                    {
                        # Referencia SOFT a la table Producto de FL_Masters
                        "prod_cod": artobj.prod_cod,
                        "prod_descripcion": artobj.descripcion,
                        "prod_unidad_medida": artobj.medidaobj.medida_cod,
                        "prod_unidad_medida_desc": artobj.medidaobj.medida,
                        "prod_pais_origen": 0,
                        "prod_pais_origen_desc": None,
                        "prod_lote": None,
                        "prod_vencimiento": None,
                        "porcentaje_iva": 10,
                        "precio_unitario": doc_voucher,
                        "precio_unitario_siniva": doc_voucher / 1.1,
                        "cantidad": 1,
                        "cantidad_devolucion": 0,
                        "exenta": 0,
                        "iva_5": 0,
                        "gravada_5": 0,
                        "iva_10": doc_voucher - (doc_voucher / 1.1),
                        "gravada_10": doc_voucher,
                        "afecto": doc_voucher,
                        "per_tipo_iva": 0,  # dPropIVA Porcentaje del gravado, cuando en un mismo ite existe g5 y exenta o g10 y exenta
                        "bonifica": False,  # Si es un descuento
                        "descuento": 0,
                        "per_descuento": 0,
                        "volumen": 0,
                        "peso": 0,
                        "observacion": "Descuento voucher",
                    }
                ],
                pagos=[
                
                    {"tipo_cod": 1, "monto": doc_total},
                ],
            )
            msg += f' con Documento Asociado {rsp_asoc.get("doc_tipo")} {rsp_asoc.get("doc_numero")} creado con exitos'

        send_task('Sifen.tasks.send_invoice', 
            kwargs={
                'username': userobj.username,
                'qdict': {
                    'dbcon': 'default',
                    'docpk': docheaderobj.pk,
                    'dattrs': to_json({
                        'full': 1,
                    })
            }
        })
        return {'success': msg, 
                'id': docheaderobj.pk, 
                'doc_numero': docheaderobj.doc_numero,
                'doc_tipo': docheaderobj.doc_tipo}

    def set_number(self, 
            timbrado,
            prof_numbers: list, 
            expd: int, 
            establecimiento: int, 
            sign_document: bool = False, 
            doc_tipo: str = 'FE'):
        eser = ekuatia_serials.Eserial()
        qdict = QueryDict(mutable=True)
        qdict.update({
            'timbrado': timbrado,
            'expd': expd,
            'establecimiento': establecimiento,
            'sign_document': sign_document,
            'ruc': self.RUC,
            'tipo': doc_tipo
        })
        logging.info(f'Get numbers for the document base on the params {qdict}')
        for prof in prof_numbers:
            qdict.update({'prof_number': prof})
        return eser.set_number(qdict=qdict)

    def release_number(self, *args, **kwargs):
        logging.info('Running release_numbert')
        qdict = kwargs.get('qdict')
        ruc = qdict.get('ruc')
        establecimiento = qdict.get('establecimiento')
        tipo = qdict.get('tipo')
        docpk = qdict.get('docpk')
        docobj = DocumentHeader.objects.get(pk=docpk)
        if docobj.lote_estado in ['RECIBIDO', 'PROCESANDO', 'CONCLUIDO']:
            return {'error': 'No es posible liberar el numero de un documento en proceso de envio'}
        if docobj.ek_estado == 'Aprobado' or \
            docobj.ek_estado == 'Rechazado' or \
            docobj.lote_estado == 'Aprobado':
            return {'error': 'No es posible liberar el numero de un documento ya procesado'}
        qed = QueryDict(mutable=True)
        qed.update({
            'timbrado': docobj.ek_timbrado,
            'establecimiento': establecimiento,
            'ruc': ruc,
            'tipo': tipo,
            'state': 'L',
            'numero': docobj.doc_numero
        })
        eser = ekuatia_serials.Eserial()
        print(qed)
        eser.set_state_numbers(qdict=qed)
        return {'success': f'Numero {docobj.doc_numero} liberado correctamente', 
                'released_numbers': qed,
                }

    def firmar_proforma(self, qsf: dict = {'ek_estado__isnull': True }):
        eser = ekuatia_serials.Eserial()
        rsp = eser.generate_pending_cdc_console(
                        pps=qsf, 
                        timbrado=self.timbradoobj.timbrado,
                        ruc=self.RUC)
        peds = rsp.get('prof_number')
        #print(rsp)
        for p in peds:
            headerobj = DocumentHeader.objects.get(prof_number=p)
            msg = f'Pedido {headerobj.prof_number} firmado correctamente en archivo {headerobj.ek_xml_file_signed}'
            logging.info(msg)

    def generar_pdf(self, prof_numbers: list):
        ek = e_kude.EKude(fl_sifen_conf.RUC)
        for prof_number in prof_numbers:
            ret = ek.eprint_doc(None, 
                        query_dict={'layout': 'full_page', 
                                    'digital': 1, 
                                    'set_name': 'factura_numero', 
                                    'prof_number': prof_number}
            )
            #print(ret)

    def get_facturas_recibo(self, *args, **kwargs):
        facturas: list = kwargs.get('facturas', [])
        fdata = []
        ndata = []
        totales = 0
        totales_nc = 0
        clobj = None
        retencion = 0
        for f in DocumentHeader.objects.filter(pk__in=facturas):
            clobj = Clientes.objects.using('fl').get(clientecodigo=f.pdv_codigo)
            if f.doc_tipo == 'NC':
                ndata.append({
                    'id': f.pk,
                    'tipo': f.doc_tipo,
                    'numero': f.get_number_full(),
                    'monto': f.get_total_venta_gs()
                })
                totales_nc += f.get_total_venta_gs()
            if f.doc_tipo == 'FE':
                retencion = f.retencionobj.retencion if f.retencionobj else 0
                fdata.append({
                    'id': f.pk,
                    'tipo': f.doc_tipo,
                    'numero': f.get_number_full(),
                    'monto': f.get_total_venta_gs(),
                    'retencion_numero': f.retencionobj.get_numero() if f.retencionobj else 0,
                    'retencion': retencion,
                })
                totales += (f.get_total_venta_gs() - retencion)
                retencion += f.retencionobj.retencion if f.retencionobj else 0
        totales = round(totales)
        return {
            'clobj': clobj,
            'ndata': ndata,
            'fdata': fdata,
            'facturas': len(fdata),
            'ncs': len(ndata),
            'total': totales,
            'retencion': retencion,
            'total_word':num2words(totales, lang='es').upper(),
            'total_nc': totales_nc,
            'total_word_nc':num2words(totales_nc, lang='es').upper()
        }
    
    def crear_recibo(self, *args: list, **kwargs: dict) -> tuple:
        today = datetime.today()
        ios = IoS()
        eser = ekuatia_serials.Eserial()
        userobj = kwargs.get('userobj')
        q: dict = kwargs.get('qdict', {})
        uc_fields: dict = from_json(q.get('uc_fields', "{}"))
        facturas = uc_fields['facturas'] 
        ncs = uc_fields['ncs'] 
        forma_pago_id = uc_fields.get('forma_pago_id')
        forma_pago = uc_fields.get('forma_pago')
        fobjs = DocumentHeader.objects.filter(pk__in=facturas)
        ncobjs = DocumentHeader.objects.filter(pk__in=ncs)
        if not fobjs:
            return {'error': 'Debe seleccionar al menos una factura'}, args, kwargs
        pfu = list(fobjs.values_list('prof_number', flat=True))
        nfu = list(ncobjs.values_list('prof_number', flat=True))
        if DocumentReciboDetail.objects.filter(prof_number__in=pfu):
            return {'error': 'Las facturas seleccionadas ya tienen recibo generado'}, args, kwargs
        if DocumentReciboDetail.objects.filter(prof_number__in=nfu):
            return {'error': 'Las notas de creditos seleccionadas ya tienen recibo generado'}, args, kwargs
        fobj = fobjs[0]
        cg = Cotizacion.objects.using('fl').order_by('id_cotizacion').last()
        if cg:
            tasa_cambio = cg.dolar_venta
        recobj = DocumentRecibo.objects.create(
            bs = fobj.bs,
            source = 'ERP',
            ext_link = 0,
            doc_moneda = fobj.doc_moneda,
            doc_fecha = today,
            doc_tipo = 'RC',
            doc_tipo_cod = 9,
            doc_tipo_desc = 'RECIBOD',
            doc_op = 'COBRO',
            doc_numero = 0,
            doc_expedicion = fobj.doc_expedicion,
            doc_establecimiento = fobj.doc_establecimiento,
            doc_establecimiento_ciudad = fobj.doc_establecimiento_ciudad,
            doc_estado = 'INICIADO',
            doc_vencimiento = arrow.get().shift(days=30).strftime('%Y-%m-%d'),
            doc_total_factura = 0,
            doc_total_nc = 0,
            doc_cobrar = 0,
            doc_retencion = 0,
            doc_efectivo = uc_fields['doc_efectivo'],
            doc_cheque = uc_fields['doc_cheque'],
            doc_cobrado = 0,
            pdv_innominado = fobj.pdv_innominado,
            pdv_pais_cod = fobj.pdv_pais_cod,
            pdv_pais = fobj.pdv_pais,
            pdv_tipocontribuyente =  fobj.pdv_tipocontribuyente,
            pdv_es_contribuyente = fobj.pdv_es_contribuyente,
            pdv_type_business = fobj.pdv_type_business,
            pdv_codigo = fobj.pdv_codigo,
            pdv_ruc = fobj.pdv_ruc,
            pdv_ruc_dv = fobj.pdv_ruc_dv,
            pdv_nombrefantasia = fobj.pdv_nombrefantasia,
            pdv_nombrefactura = fobj.pdv_nombrefactura,
            pdv_direccion_entrega = fobj.pdv_direccion_entrega,
            pdv_dir_calle_sec = fobj.pdv_dir_calle_sec,
            pdv_direccion_comple = fobj.pdv_direccion_comple,
            pdv_numero_casa = fobj.pdv_numero_casa,
            pdv_numero_casa_entrega = fobj.pdv_numero_casa_entrega,
            pdv_dpto_cod = fobj.pdv_dpto_cod,
            pdv_dpto_nombre = fobj.pdv_dpto_nombre,
            pdv_distrito_cod = fobj.pdv_distrito_cod,
            pdv_distrito_nombre = fobj.pdv_distrito_nombre,
            pdv_ciudad_cod = fobj.pdv_ciudad_cod,
            pdv_ciudad_nombre = fobj.pdv_ciudad_nombre,
            pdv_telefono = fobj.pdv_telefono,
            pdv_celular = fobj.pdv_celular,
            pdv_email = fobj.pdv_email,
            tasa_cambio = tasa_cambio,
            cargado_fecha = today,
            cargado_usuario = userobj.username,
        )
        retencion = 0
        for fobj in fobjs:
            retencion += fobj.retencionobj.retencion if fobj.retencionobj else 0
            retencion_numero = fobj.retencionobj.retencion_numero if fobj.retencionobj else 0
            recobj.documentrecibodetail_set.create(
                tipo = fobj.doc_tipo,
                prof_number = fobj.prof_number,
                establecimiento = fobj.doc_establecimiento,
                expedicion = fobj.doc_expedicion,
                numero = fobj.doc_numero,
                cobrado = fobj.get_total_venta_gs(),
                total = fobj.get_total_venta_gs(),
                saldo = 0,
                retencion_numero = retencion_numero,
                retencion = retencion,
                observacion = 'ND',
                cargado_fecha = today,
                cargado_usuario = userobj.username
            )
        for ncobj in ncobjs:
            recobj.documentrecibodetail_set.create(
                tipo = ncobj.doc_tipo,
                prof_number = ncobj.prof_number,
                establecimiento = ncobj.doc_establecimiento,
                expedicion = ncobj.doc_expedicion,
                numero = ncobj.doc_numero,
                cobrado = ncobj.get_total_venta_gs(),
                total = ncobj.get_total_venta_gs(),
                saldo = 0,
                observacion = 'ND',
                cargado_fecha = today,
                cargado_usuario = userobj.username
            )
        qdict = QueryDict(mutable=True)
        qdict.update({
            'establecimiento': fobj.doc_establecimiento,
            'prof_number': fobj.prof_number
        })
        eser.set_number_receive(qdict=qdict)
        #Actualizamos las facturas
        fobjs.update(doc_estado='CONCLUIDO', doc_saldo=0, forma_pago_id=forma_pago_id, forma_pago=forma_pago)
        ncobjs.update(doc_estado='CONCLUIDO', doc_saldo=0)
        #actualizamos los totales
        total_fe = recobj.documentrecibodetail_set.filter(tipo='FE').aggregate(total=Sum('total')).get('total')
        total_nc = recobj.documentrecibodetail_set.filter(tipo='NC').aggregate(total=Sum('total')).get('total')
        if not total_fe: total_fe = 0
        if not total_nc: total_nc = 0
        #recobj.doc_cobrado = total_fe
        recobj.doc_retencion = retencion
        recobj.doc_cobrado = retencion+uc_fields['doc_efectivo']+uc_fields['doc_cheque']
        recobj.doc_total_nc = total_nc
        recobj.doc_total_factura = total_fe
        recobj.save()
        return {'success': 'Recibo creado'}, args, kwargs

    def generando_documentrecibo(self, *args: list, **kwargs: dict) -> tuple:
        userobj = kwargs.get('userobj')
        q: dict = kwargs.get('qdict', {})
        dbcon = q.get('dbcon')
        id = q.get('id')
        self.crear_documentrecibo_pdf(
            username=userobj.username,
            userobj=userobj,
            qdict={
                'dbcon': dbcon,
                'id': id,
                'dattrs': to_json({
                    'media_path': True
                })
        })
        docobj = DocumentRecibo.objects.using(dbcon).get(pk=id)
        return {'success': f'Creacion PDF', 'documentrecibo_urls': docobj.documentrecibo_urls() }
    
    @tracktask
    def crear_documentrecibo_pdf(self, *args: list, **kwargs: dict) -> dict:
        q: dict = kwargs.get('qdict')
        dbcon = q.get('dbcon')
        id =  q.get('id')
        dattrs = q.get('dattrs')
        docobj = DocumentRecibo.objects.using(dbcon).get(id=id)
        mq = QueryDict(mutable=True)
        fv = {
            'tmpl': 'Sifen/DocumentReciboRptUi.html',
            'model_app_name':'Sifen',
            'model_name':'DocumentRecibo',
            'pk':id,
            'dattrs': dattrs,
            'dbcon':'default',
            'surround': 'BaseAmReportUi.html',
            'g_pdf': 0,
            'g_pdf_kit': 1,
            'page-size': 'A5',
            'orientation': 'Portrait',
            'page-height': '250.000212',
            'page-width': '210.000088',
            'margin-top': '0',
            'margin-right': '0',
            'margin-bottom': '0',
            'margin-left': '0',
            'no-outline': None
        }
        mq.update(fv)
        mrq = HttpRequest()
        mrq.user = kwargs.get('userobj')
        mrq.method = 'GET'
        mrq.GET = mq
        io_rpt = IoRpt()
        rp = io_rpt.rpt_view(mrq)
        pdf_file = rp.get('pdf_file')
        with open(pdf_file, 'rb') as f:
            docobj.pdf_file = File(f, name=os.path.basename(pdf_file))
            docobj.save()
        return {
            'success': f'Pdf del documento {docobj.id} creado con exito',
            'pdf_file': pdf_file,
            'html_file': rp.get('html_file')
        }
    
    def search_factura_relacionada(self,*args, **kwargs):
        qf = kwargs.get('qf')
        # Get prof_numbers of invoices already used in NC documents
        used_prof_numbers = DocumentHeader.objects.filter(
            doc_tipo='NC'
        ).values_list('doc_loop_link', flat=True).distinct()

        # Exclude invoices that are already used in NC
        return DocumentHeader.objects.filter(**qf).exclude(
            prof_number__in=used_prof_numbers
        ).order_by('-doc_numero')[0:100]
    
    def search_factura_retencion(self, *args, **kwargs):
        qf = kwargs.get('qf')
        return DocumentHeader.objects.filter(**qf).order_by('-doc_numero')[0:10]

    def rpt_libro_vta(self, *args: list, **kwargs: dict) -> dict:
        fog = Font(name='Verdana',
                     size=7,
                     bold=False,
                     italic=False,
                     vertAlign=None,
                     underline='none',
                     strike=False,
                     )    
        border_style = 'thin'
        l_border = Border(
            left=Side(border_style=border_style, color="000000"),
        )
        r_border = Border(
            right=Side(border_style=border_style, color="000000"),
        )
        b_border = Border(
            bottom=Side(border_style=border_style, color="000000"),
        )
        t_border = Border(
            top=Side(border_style=border_style, color="000000"),
        )
        lb_border = Border(
            left=Side(border_style=border_style, color="000000"),
            bottom=Side(border_style=border_style, color="000000")
        )

        rb_border = Border(
            right=Side(border_style=border_style, color="000000"),
            bottom=Side(border_style=border_style, color="000000")
        )
        lt_border = Border(
            left=Side(border_style=border_style, color="000000"),
            top=Side(border_style=border_style, color="000000")
        )

        rt_border = Border(
            right=Side(border_style=border_style, color="000000"),
            top=Side(border_style=border_style, color="000000")
        )
        lr_border = Border(
            left=Side(border_style=border_style, color="000000"),
            right=Side(border_style=border_style, color="000000"),
            
        )
        all_border = Border(
            left=Side(border_style=border_style, color="000000"),
            right=Side(border_style=border_style, color="000000"),
            top=Side(border_style=border_style, color="000000"),
            bottom=Side(border_style=border_style, color="000000")
        )
        ios = IoS()
        userobj = kwargs.get('userobj')
        q: dict = kwargs.get('qdict', {})
        dbcon = q.get('dbcon')
        #ejercicio = q.get('ejercicio')
        #periodo = q.get('periodo')
        qs_doc_fecha_0 = q.get('qs_doc_fecha_0')
        qs_doc_fecha_1 = q.get('qs_doc_fecha_1')

        f_desce = arrow.get(qs_doc_fecha_0).format('DD/MM/YY')
        f_hasta = arrow.get(qs_doc_fecha_1).format('DD/MM/YY')
        
        ndata = []
        sname = f'ventas_{qs_doc_fecha_0}_{qs_doc_fecha_1}'.replace('-', '')
        wb = Workbook()
        razon_social = self.bsobj.nombrefactura

        left_alignment = io_styles.left_alignment()
        right_alignment = io_styles.right_alignment()
        center_alignment = io_styles.center_alignment()
        title_font = io_styles.title_font(font='Verdana', size=7)
        title_fill = io_styles.title_fill(color="D8D8D8")
        swei = wb.active
        swei.title = sname
        swei.merge_cells('A5:K5')
        
        swei['A5'] = f'LIBRO VENTAS DEL PERIODO {f_desce} AL {f_hasta}'
        swei['A5'].font = title_font
        swei['A5'].fill = title_fill
        swei['A5'].alignment = center_alignment
        swei.merge_cells('A2:C2')
        razon_social = 'FRONTLINER S.A.'
        swei['A2'] = f'Razon Social: {razon_social}'
        #swei.merge_cells('A3:C3')
        #swei['B2'] = razon_social
        swei.merge_cells('A3:C3')
        ruc = f'{self.bsobj.ruc}-{self.bsobj.ruc_dv}'
        ruc = '80070523-8'
        swei['A3'] = f'Ruc: {ruc}'
        swei.merge_cells('A4:F4')
        swei['A4'] = f'Direccion: {self.bsobj.direccion}'
        #swei.merge_cells('A5:B5')
        #swei['B4'] = self.bsobj.direccion

        
        swei = io_styles.set_sheet_auto_width(swei, start_row=1)
        titles = [
            'A6', 'B6', 'C6',
            'D6', 'F6', 'G6',
            'I6', 'J6'
        ]

        swei.merge_cells('A7:A8')
        swei.merge_cells('B7:B8')
        swei.merge_cells('K7:K8')
        swei.merge_cells('L7:L8')
        swei.merge_cells('C7:D7')
        swei.merge_cells('E7:J7')

        header_text = [
            'A7', 'B7', 'C7', 'D7',
            'E7', 'F7', 'G7', 'H7', 'I7',
            'J7', 'K7','L7'
        ]
        for tt in titles:
            swei[tt].font = title_font
            swei[tt].alignment = left_alignment

        for tt in header_text:
            swei[tt].font = fog
            swei[tt].alignment = center_alignment
        swei.sheet_view.showGridLines = False
        tc = {
            'A7': {'col': 'Dia',  'start': 'A7', 'width': 5.62},
            'B7': {'col': 'Documento Numero',  'start': 'B7', 'width': 10.15},
            'C7': {'col': 'Cliente de Bienes y Servic',  'start': 'C7', 'width': 23.90},
            'C8': {'col': 'R.Soc/Apell./Nomb',  'start': 'C8', 'width': 19.31},
            'D8': {'col': 'RUC',  'start': 'D8', 'width': 10.31},
            'E7': {'col': 'Valores de ventas y servicios',  'start': 'E7', 'width': 40.50},
            'E8': {'col': 'Grav 10%',  'start': 'E8', 'width': 12.69},
            'F8': {'col': 'Grav 5%',  'start': 'F8', 'width': 12.69},
            'G8': {'col': 'Iva 10%',  'start': 'G8', 'width': 12.69},
            'H8': {'col': 'Iva 5%',  'start': 'H8', 'width': 12.69},
            'I8': {'col': 'Exento',  'start': 'I8', 'width': 12.69},
            'J8': {'col': 'Total',  'start': 'J8', 'width': 12.69},
            'K7': {'col': 'Redondeo',  'start': 'K7', 'width': 12.69},
            'L7': {'col': 'Retencion',  'start': 'L7', 'width': 12.69},
        }
        for sc, t in tc.items():
            cf = t.get('start')
            swei[cf] = t.get('col')
            swei[cf].font = fog
            swei[cf].alignment = left_alignment
            swei.column_dimensions[sc[0]].width = t.get('width') 

        qs_doc_tipo = q.get('qs_doc_tipo')
        qs_cobranza = q.get('qs_cobranza')
        qs_clientecodigo_id = q.get('qs_clientecodigo_id')

        pps = {
            'doc_fecha__range': (qs_doc_fecha_0, qs_doc_fecha_1),
            # 'doc_fecha__year': ejercicio, 
            # 'doc_fecha__month': periodo, 
            'doc_tipo__in': ['FE', 'AF']
        }
        if qs_doc_tipo:
            pps.pop('doc_tipo__in', None)
            pps['doc_tipo'] = qs_doc_tipo
        if qs_cobranza:
            if qs_cobranza == 'pendiente_cobro':
                pps['doc_saldo__gt'] = 0
            else:
                pps['doc_saldo'] = 0
        if qs_clientecodigo_id:
            pps['pdv_codigo'] = qs_clientecodigo_id
        for docobj in DocumentHeader.objects.filter(
                        **pps
                ).order_by('doc_numero', 'doc_fecha'):
            ndata.append({
                'dia': docobj.doc_fecha.day,
                'doc_numero': docobj.get_number_full(),
                'pdv_cliente': docobj.pdv_nombrefactura,
                'pdv_codigo': docobj.pdv_ruc,
                'gravada_10': docobj.get_total_gravada_10(),
                'gravada_5': docobj.get_total_gravada_5(),
                'iva_10': docobj.get_ivas_10_master(),
                'iva_5': docobj.get_ivas_5_master(),
                'exento': docobj.get_total_exenta(),
                'total': docobj.get_total_operacion_redondeo(),
                'redondeo': docobj.doc_redondeo,
                'retencion': docobj.retencionobj.retencion if docobj.retencionobj else 0
                
            })
        if not ndata:
            return {'error': f'Sin datos en el periodo {f_desce} al {f_hasta}'}
        ncols = [
            'dia','doc_numero','pdv_cliente',
            'pdv_codigo','gravada_10','gravada_5',
            'iva_10','iva_5','exento',
            'total','redondeo','retencion',
        ]
        dfp = pd.DataFrame(ndata)
        rows = dataframe_to_rows(dfp[ncols], index=False, header=False)
        r_idx = 0
        for r_idx, row in enumerate(rows, 9):
            for c_idx, value in enumerate(row, 1):
                # if c_idx == 8:
                #     value = round(value, 3)                
                # if c_idx == 9:
                #     value = round(value, 2)
                d = swei.cell(row=r_idx, column=c_idx, value=value)
                d.font = fog
                d.alignment = left_alignment
                # if c_idx in [1, 3,4]:
                #     d.alignment = center_alignment
                # if c_idx in [7, 8, 9]:
                #     d.alignment = right_alignment
                d.border = all_border
        lrow = r_idx
        
        # tcs = [
        #     ('A{}:B{}', 'Master Airline:'),
        #     ('A{}:B{}', 'Fecha Cierre:'),
        #     ('A{}:B{}', 'Number of Bags:'),
        #     ('A{}:B{}', 'TOTAL KG:'),
        #     ('A{}:B{}', 'Fecha CBM:'),

        # ]
        # for tx, tc in enumerate(tcs):
        #     lrow += 1
        #     tr = tc[0].format(lrow, lrow)
        #     swei.merge_cells(tr)
        #     swei[tr.split(':')[0]] = tc[1]
        # lrow += 1
        # swei.merge_cells(f'A{lrow}:I{lrow}')
        # swei[f'A{lrow}'] = 'Comment:'
        # lrow += 1
        # swei.merge_cells(f'A{lrow}:I{lrow+2}')

        # swei[f'A{lrow}'] = f"""I, Diego Lopez, from {embobj.codoficina.nomoficina}, today, {MONTH} {DAY}st, hereby certify and confirm that ALL shipments prepared by {razon_social} are screened 100% and does not contains any dangerous goods or hazmat materials, nor any restricted or prohibited cargo for air shipping purposes as per IATA, ICAO, OACI or any other domestic and international regulations."""
        

        # lrow += 3
        # swei.merge_cells(f'A{lrow}:I{lrow}')

        # swei[f'A{lrow}'] = 'We also declare hereby, that ALL of our shipments are not subject to any export license processes and/or restrictions and they are declare as "EEI NO REQUIRED as per SEC. 30.37 (a) FTSR - EAR99 - NLR".'
        # lrow += 1
        # swei.merge_cells(f'A{lrow}:E{lrow}')
        # swei[f'A{lrow}'] = 'Following the current IATA ICAO DG regulations, I hereby:'
        # lrow += 1
        # swei[f'A{lrow}'] = '(Check one)'
        # lrow += 2
        # swei[f'A{lrow}'] = 'YES'
        # swei.merge_cells(f'B{lrow}:I{lrow}')
        # swei[f'B{lrow}'] = 'CONFIRM this freight contains any kind of Lithium Batteries, in equipment or not. Cargo has been labeled accordingly and under compliance.'
        # lrow += 1
        # swei.merge_cells(f'A{lrow}:I{lrow+2}')
        # swei[f'A{lrow}'] = f"""I Diego Lopez on behalf of {razon_social}, and all its employees and subsidiaries, confirm we will not held Ground Cargo Transportation INC or any of its affiliates responsible or liable of any penalties, fines, damages or loss incur by any of our shipments. The shipper agrees to indemnity Ground Cargo Transportation INC for any and all costs, fees and expenses Ground Cargo Transportation INC incurs as a result of the shipper's violation of any local, state or federal laws or regulations or from tendering any prohibited item for shipment."""

        #Borders
        borders  = OrderedDict()
        # borders['bottom'] = ['A3:H3', 'A10:H10', 'F7:F7']
        # borders['top'] = ['A11:I11', 'F8:F8']
        # borders['right'] = ['C2:C3', 'E2:E3', 
        #                     'I2:I3', 
        #                     'D4:D10', 'B4:B10']
        # borders['left'] = ['G4:G7','F8:F10','G8:G10', 'J4:J10']
        # borders['left_bottom'] = ['F7:I7']
        borders['all'] = ['A7:L8']
        
        # if r_idx > 0:
        #     lrow += 2
        #     borders['all'].append(f'A{r_idx}:I{lrow}')
        #     for row in swei[f'A{r_idx}:I{lrow}']:
        #         for cell in row:
        #             cell.font = fog
        #             cell.alignment = left_alignment

        for idx, b in borders.items():
            if idx == 'all':
                for c_bo in b:
                    for row in swei[c_bo]:
                        for cell in row:
                            cell.border = all_border
            if idx == 'left':
                for c_bo in b:
                    for row in swei[c_bo]:
                        for cell in row:
                            cell.border = l_border
            if idx == 'right':
                for c_bo in b:
                    for row in swei[c_bo]:
                        for cell in row:
                            cell.border = r_border
            if idx == 'bottom':
                for c_bo in b:
                    for row in swei[c_bo]:
                        for cell in row:
                            cell.border = b_border
            if idx == 'top':
                for c_bo in b:
                    for row in swei[c_bo]:
                        for cell in row:
                            cell.border = t_border
            if idx == 'right_bottom':
                for c_bo in b:
                    for row in swei[c_bo]:
                        for cell in row:
                            cell.border = rb_border
            if idx == 'right_top':
                for c_bo in b:
                    for row in swei[c_bo]:
                        for cell in row:
                            cell.border = rt_border
            if idx == 'left_bottom':
                for c_bo in b:
                    for row in swei[c_bo]:
                        for cell in row:
                            cell.border = lb_border
            if idx == 'left_top':
                for c_bo in b:
                    for row in swei[c_bo]:
                        for cell in row:
                            cell.border = lt_border
        fpath = f'{settings.MEDIA_ROOT}/tmp/{sname}.xls'
        furl = f'media/tmp/{sname}.xls'
        wb.save(fpath)
        return {'success': 'Hecho', 'furl': furl, 'file_name': sname, 'full_path': fpath }
    
    
    def crear_retencion(self, *args: list, **kwargs: dict) -> tuple:
        gdata = mng_gmdata.Gdata()
        ios = IoS()
        userobj = kwargs.get('userobj')
        q: dict = kwargs.get('qdict', {})
        dbcon = q.get('dbcon')
        uc_fields: dict = from_json(q.get('uc_fields', "{}"))
        uc_fields['retencion_numero'] = uc_fields['retencion_numero'].replace('-', '').replace('_', '')
        if len(uc_fields['retencion_numero']) < 13:
            return {'error': 'El numero de la retencion es invalido'}, args, kwargs
        uc_fields['cargado_usuario'] = userobj.first_name
        uc_fields['cargado_fecha'] = datetime.now()
        ff = ios.form_model_fields(uc_fields, Retencion._meta.fields)
        for rr in ff:
            uc_fields.pop(rr)
        rnorm, rrm, rbol = ios.format_data_for_db(Retencion, uc_fields)
        for c in rrm: uc_fields.pop(c)
        for f, fv in rbol: uc_fields[f] = fv
        for f, fv in rnorm: uc_fields[f] = fv
        clobj = Clientes.objects.using('fl').get(clientecodigo=uc_fields.get('pdv_codigo'))
        if Retencion.objects.using(dbcon).filter(retencion_numero=uc_fields['retencion_numero']):
            return {'error': f'El numero de retencion {uc_fields["retencion_numero"]} ya fue cargado para el cliente {clobj.normalize_name()}'}, args, kwargs
        uc_fields['pdv_ruc'] = clobj.ruc
        uc_fields['pdv_ruc_dv'] = gdata.calculate_dv(clobj.ruc)
        uc_fields['pdv_nombrefantasia'] = clobj.normalize_name()
        uc_fields['pdv_nombrefactura'] = clobj.razon_social
        docobj = DocumentHeader.objects.get(pk=uc_fields['doc_loop_link'])
        uc_fields['doc_relacion_numero'] = docobj.doc_numero
        uc_fields['doc_relacion_monto'] = docobj.doc_total_redondeo
        if float(uc_fields['retencion']) > docobj.doc_total_redondeo:
            return {'error': 'El valor de la retencion no puede ser superior al de la factura'}, args, kwargs
        
        retobj = Retencion.objects.using(dbcon).create(**uc_fields)
        docobj.retencionobj = retobj
        docobj.save()
        files: dict = kwargs.get('files')
        for name, fobj in files.items():
            fname = f'op_{fobj.name}'
            dfobj = File(fobj, name=fname)
            setattr(retobj, name, dfobj)
        retobj.save()
        return {'success': f'Retencion creada', 'record_id': retobj.id}, args, kwargs
    
    def actualizar_retencion(self, *args: list, **kwargs: dict) -> tuple:
        ios = IoS()
        userobj = kwargs.get('userobj')
        q: dict = kwargs.get('qdict', {})
        dbcon = q.get('dbcon')
        files: dict = kwargs.get('files')
        uc_fields: dict = from_json(q.get('uc_fields', "{}"))
        uc_fields['retencion_numero'] = uc_fields['retencion_numero'].replace('-', '').replace('_', '')
        ff = ios.form_model_fields(uc_fields, Retencion._meta.fields)
        for rr in ff:
            uc_fields.pop(rr)
        rnorm, rrm, rbol = ios.format_data_for_db(Retencion, uc_fields, update=True)
        for c in rrm: uc_fields.pop(c)
        for f, fv in rbol: uc_fields[f] = fv
        for f, fv in rnorm: uc_fields[f] = fv
        uc_fields['actualizado_usuario'] = userobj.first_name
        uc_fields['actualizado_fecha'] = datetime.now()
        retobj = Retencion.objects.using(dbcon).get(id=uc_fields.get('id'))
        if retobj.doc_loop_link != int(uc_fields['doc_loop_link']):
            docobj = DocumentHeader.objects.get(pk=retobj.doc_loop_link)
            docobj.retencionobj = None
            docobj.save()
            docobj = DocumentHeader.objects.get(pk=uc_fields['doc_loop_link'])
            uc_fields['doc_loop_link'] = docobj.pk
            uc_fields['doc_relacion_numero'] = docobj.doc_numero
            uc_fields['doc_relacion_monto'] = docobj.doc_total_redondeo
            docobj.retencionobj = retobj
            docobj.save()
            if float(uc_fields['retencion']) > docobj.doc_total_redondeo:
                return {'error': 'El valor de la retencion no puede ser superior al de la factura'}, args, kwargs
        u_fields = ios.get_differences_fields(model_to_dict(retobj), uc_fields)
        files: dict = kwargs.get('files')
        if files:
            for name, fobj in files.items():
                fname = f'order_{fobj.name}'
                dfobj = File(fobj, name=fname)
                setattr(retobj, name, dfobj)
            retobj.save()
        if u_fields:
            Retencion.objects.using(dbcon).filter(id=retobj.id).update(**u_fields)
        if not u_fields:
            return {'info': f'Nada que actualizar'}, args, kwargs
        return {'success': f'Pedido de compra {retobj.id} actualizado con exitos', 'record_id': retobj.id }, args, kwargs
    
    def anular_documento(self, *args, **kwargs) -> dict:
        today = date.today()
        userobj = kwargs.get('userobj')
        q: dict = kwargs.get('qdict', {})
        facturas = q.getlist('facturas')
        for docobj in DocumentHeader.objects.filter(pk__in=facturas, anulado_fecha__isnull=True):
            dd = (today - docobj.doc_fecha).days
            if (dd > 3):
                #Generar nc
                qdict = QueryDict(mutable=True)
                th = {
                    "pdv_codigo": docobj.pdv_codigo,
                    "pdv_ruc": docobj.pdv_ruc,
                    "pdv_ruc_dv": docobj.pdv_ruc_dv,
                    "pdv_email": docobj.pdv_email,
                    "pdv_direccion_entrega": docobj.pdv_direccion_entrega,
                    "pdv_nombrefactura": docobj.pdv_nombrefactura,
                    "doc_cre_tipo_cod": "2",
                    "doc_moneda": docobj.doc_moneda,
                    "doc_fecha": today.strftime('%Y-%m-%d'),
                    "doc_tipo": "NC",
                    "doc_numero": 0,
                    "doc_vencimiento": today.strftime('%Y-%m-%d'),
                    "tasa_cambio": docobj.tasa_cambio,
                    "doc_relacion_cdc": docobj.pk,
                    "doc_motivo": "Crédito incobrable",
                    "observacion": "Anulado",
                    "detail": []
                }
                for ddobj in docobj.documentdetail_set.filter(anulado=False, bonifica=False):
                    th["detail"].append(
                        {
                            "doc_moneda": docobj.doc_moneda,
                            #"moneda": ddobj.moneda,
                            "prod_cod": ddobj.prod_cod,
                            "prod_descripcion": ddobj.prod_descripcion,
                            "precio_unitario_source": ddobj.precio_unitario_source,
                            "precio_unitario": ddobj.precio_unitario,
                            "precio_unitario_siniva": ddobj.precio_unitario_siniva,
                            "cantidad": ddobj.cantidad,
                            "exenta": float(ddobj.exenta),
                            "iva_5": float(ddobj.iva_5),
                            "gravada_5": float(ddobj.gravada_5),
                            "gravada_10": float(ddobj.gravada_10),
                            "iva_10": float(ddobj.iva_10),
                            "afecto": float(ddobj.afecto),
                            #"total": ddobj.total,
                        }
                    )
                qdict.update({'uc_fields': to_json(th)})
                qdict.update({'files': {}})
                rsp = self.crear_documentheader(
                    userobj=userobj,
                    qdict=qdict,
                    files={}
                )
                if rsp[0].get('error'): continue
                docobj.anulado_tipo = "NC"
            else:
                docobj.anulado_tipo = "SIFEN"
            docobj.anulado_fecha = datetime.now()
            docobj.anulado_usuario = userobj.first_name
            docobj.save()
        return {'success': 'Facturas anuladas'}
    
    def search_retencion_libre(self, *args, **kwargs):
        qf = kwargs.get('qf')
        rets = list(DocumentReciboDetail.objects.filter(
                    recobj__pdv_codigo=qf.get('pdv_codigo'),
                    tipo='RETENCION'
            ).values_list('numero', flat=True))
        print(qf, rets)
        return Retencion.objects.filter(**qf)\
                .exclude(retencion_numero__in=rets)\
                .order_by('-retencion_numero')[0:10]
    
    def metodos_de_pagos(self, *args, **kwargs) -> list:
        q: dict = kwargs.get('qdict', {})
        pp = MetodosPago.objects.all().values()
        if q.get('doc_cre_tipo_cod'):
            if int(q.get('doc_cre_tipo_cod')) == 1:
                return list(filter(lambda x: x.get('es_factura_credito') == False , pp))
            if int(q.get('doc_cre_tipo_cod')) == 2:
                return list(filter(lambda x: x.get('es_factura_credito') == True , pp))
        return pp

    def create_documentheader(self, *args, **kwargs) -> tuple:
        """Create DocumentHeader record - Manual invoice creation approach"""
        ios = IoS()
        eser = ekuatia_serials.Eserial()
        gdata = mng_gmdata.Gdata()
        userobj = kwargs.get('userobj')
        q: dict = kwargs.get('qdict', {})
        dbcon = q.get('dbcon', 'default')
        uc_fields: dict = from_json(q.get('uc_fields', {}))
        send_save = uc_fields.pop('send_save', False)

        pk = uc_fields.get('id')

        # Extract details from uc_fields
        details = uc_fields.pop('details', [])

        if not details:
            return {'error': 'Faltan los detalles del documento'}, args, kwargs

        # Get timbrado from form
        if not pk:
            # Set default values
            uc_fields['cargado_usuario'] = userobj.first_name if userobj else 'system'
            uc_fields['cargado_fecha'] = datetime.now(tz=ZoneInfo('America/Asuncion'))
            uc_fields['doc_fecha'] = datetime.now(tz=ZoneInfo('America/Asuncion')).date()

        timbrado_id = uc_fields.pop('timbrado_id', None)
        if timbrado_id:
            timbradoobj = Etimbrado.objects.get(pk=timbrado_id)
        else:
            timbradoobj = self.timbradoobj

        # Get establecimiento from form
        doc_establecimiento = uc_fields.get('doc_establecimiento')
        if not doc_establecimiento:
            estab_obj = Eestablecimiento.objects.filter(timbradoobj=timbradoobj).first()
            if not estab_obj:
                return {'error': 'No hay establecimiento configurado'}, args, kwargs
            doc_establecimiento = estab_obj.establecimiento

        # Calculate RUC DV if not provided
        if not uc_fields.get('pdv_ruc_dv'):
            uc_fields['pdv_ruc_dv'] = gdata.calculate_dv(uc_fields.get('pdv_ruc', ''))

        # Set pdv_nombrefantasia to pdv_nombrefactura if not provided
        if not uc_fields.get('pdv_nombrefantasia'):
            uc_fields['pdv_nombrefantasia'] = uc_fields.get('pdv_nombrefactura', '')

        if uc_fields.get('pdv_tipocontribuyente') == '3':
            uc_fields['pdv_es_contribuyente'] = False
            uc_fields['pdv_innominado'] = True
            
            
        
        #uc_fields['doc_op'] = 'VTA'
        uc_fields['doc_estado'] = 'CREADO'
        uc_fields['source'] = 'MANUAL'
        uc_fields['ext_link'] = 0

        # Business info
        uc_fields['bs'] = self.bsobj.name
        uc_fields['ek_bs_ruc'] = self.bsobj.ruc
        uc_fields['ek_bs_ae'] = self.bsobj.actividadecoobj.nombre_actividad
        uc_fields['ek_bs_ae_cod'] = self.bsobj.actividadecoobj.codigo_actividad
        uc_fields['ek_idcsc'] = timbradoobj.fcsc
        uc_fields['ek_idscsc'] = timbradoobj.scsc
        uc_fields['ek_timbrado'] = timbradoobj.timbrado
        uc_fields['ek_timbrado_vigencia'] = timbradoobj.inicio
        uc_fields['ek_timbrado_vencimiento'] = timbradoobj.vencimiento

        # Establecimiento info
        uc_fields['doc_establecimiento'] = doc_establecimiento
        uc_fields['doc_establecimiento_ciudad'] = uc_fields.get('doc_establecimiento_ciudad', 'ASUNCIÓN')
        uc_fields['doc_expedicion'] = uc_fields.get('doc_expedicion', 1)

        # PDV defaults
        uc_fields['pdv_pais_cod'] = uc_fields.get('pdv_pais_cod', 'PRY')
        uc_fields['pdv_pais'] = uc_fields.get('pdv_pais', 'Paraguay')
        if (uc_fields.get('pdv_ruc_dv') is None) or (uc_fields.get('pdv_ruc_dv') == ''):
            uc_fields['pdv_tipocontribuyente'] = 1
            uc_fields['pdv_es_contribuyente'] = False
        uc_fields['pdv_type_business'] = uc_fields.get('pdv_type_business', 'B2B')
        if not pk:
            uc_fields.pop('doc_numero', None)
        # Operation type defaults
        uc_fields['doc_tipo_ope'] = uc_fields.get('doc_tipo_ope', 2)
        uc_fields['doc_tipo_ope_desc'] = uc_fields.get('doc_tipo_ope_desc', 'Prestación de servicios')
        uc_fields['doc_op_pres_cod'] = uc_fields.get('doc_op_pres_cod', 1)
        uc_fields['doc_op_pres'] = uc_fields.get('doc_op_pres', 'Operación presencial')

        # Credit type - get description from code
        doc_cre_tipo_cod = int(uc_fields.get('doc_cre_tipo_cod', 1))

        if doc_cre_tipo_cod == 2 and not uc_fields.get('doc_vencimiento'):
            return {'error': 'Debe especificar los días de crédito para facturas a crédito'}, args, kwargs
        if doc_cre_tipo_cod == 1:
            # Vencimiento
            uc_fields['doc_vencimiento'] = arrow.get().shift(days=30).strftime('%Y-%m-%d')

        # Convertir doc_vencimiento a date si viene como string
        if uc_fields.get('doc_vencimiento'):
            if isinstance(uc_fields['doc_vencimiento'], str):
                try:
                    uc_fields['doc_vencimiento'] = datetime.strptime(uc_fields['doc_vencimiento'], '%Y-%m-%d').date()
                except ValueError:
                    return {'error': 'Formato de fecha de vencimiento inválido. Use YYYY-MM-DD'}, args, kwargs

        uc_fields['doc_cre_tipo_cod'] = doc_cre_tipo_cod
        uc_fields['doc_cre_tipo'] = fl_sifen_conf.K_CRE_TIPO_COD.get(doc_cre_tipo_cod, 'Contado')

        # Payment type defaults
        uc_fields['doc_tipo_pago_cod'] = uc_fields.get('doc_tipo_pago_cod', 1)
        uc_fields['doc_tipo_pago'] = uc_fields.get('doc_tipo_pago', fl_sifen_conf.K_TIPO_PAGO.get(1, 'Efectivo'))

        

        # Initialize totals
        uc_fields['doc_total'] = 0
        uc_fields['doc_iva'] = 0
        uc_fields['doc_exenta'] = 0
        uc_fields['doc_g10'] = 0
        uc_fields['doc_i10'] = 0
        uc_fields['doc_g5'] = 0
        uc_fields['doc_i5'] = 0
        uc_fields['doc_descuento'] = 0
        uc_fields['doc_per_descuento'] = 0
        # Preserve doc_descuento_global from frontend if provided
        if 'doc_descuento_global' not in uc_fields or uc_fields.get('doc_descuento_global') in [None, '']:
            uc_fields['doc_descuento_global'] = 0
        else:
            uc_fields['doc_descuento_global'] = float(uc_fields['doc_descuento_global'])
        uc_fields['doc_saldo'] = 0
        uc_fields['doc_pago'] = 0
        uc_fields['doc_costo'] = 0
        uc_fields['doc_redondeo'] = 0
        uc_fields['peso'] = 0
        uc_fields['volumen'] = 0
        uc_fields['observacion'] = uc_fields.get('observacion', '').strip() if uc_fields.get('observacion') else ''

        # Exchange rate - get from Cotizacion or default to 1
        if uc_fields.get('doc_moneda') == 'USD':
            cot = Cotizacion.objects.order_by('-id').first()
            uc_fields['tasa_cambio'] = cot.venta if cot else Decimal('7500')
        else:
            uc_fields['tasa_cambio'] = Decimal('1')

        # Clean fields not in model
        if uc_fields.get('doc_numero') and pk:
            uc_fields.pop('doc_numero', None)
        rnorm, rrm, rbol = ios.format_data_for_db(DocumentHeader, uc_fields)
        for c in rrm: uc_fields.pop(c, None)
        for f, fv in rbol: uc_fields[f] = fv
        for f, fv in rnorm: uc_fields[f] = fv
        ff = ios.form_model_fields(uc_fields, DocumentHeader._meta.fields)
        for rr in ff:
            uc_fields.pop(rr, None)

        # Create DocumentHeader
        if pk:
            docobj = DocumentHeader.objects.using(dbcon).get(pk=pk)
            if docobj.lote_estado in ['RECIBIDO', 'PROCESANDO', 'CONCLUIDO']:
                return {'error': f'No se pude modificar con el lote en estado {docobj.lote_estado}'}, args, kwargs
            if docobj.lote_msg == 'Aprobado':
                if not docobj.ek_estado:
                    docobj.ek_estado = 'Aprobado'
                    docobj.save()
                return {'error': 'No se puede modificar un documento ya aprobado por el SIFEN'}, args, kwargs
            if int(uc_fields['pdv_tipocontribuyente']) == 2:
                uc_fields['pdv_es_contribuyente'] = True
            DocumentHeader.objects.using(dbcon).filter(pk=pk).update(**uc_fields)
            docobj = DocumentHeader.objects.using(dbcon).get(pk=pk)
            docobj.documentdetail_set.all().delete()
        else:
            docobj = DocumentHeader.objects.using(dbcon).create(**uc_fields)



        # Process details
        doc_redondeo = 0
        for d in details:
            precio_unitario = float(str(d.get('precio_unitario', 0)))
            cantidad = float(str(d.get('cantidad', 1)))

            # Check if product is from Producto model or manual entry
            prod_cod = d.get('prod_cod')

            try:
                prodobj = Producto.objects.get(prod_cod=prod_cod)
            except Producto.DoesNotExist:
                prod_autocreado = True
                # If not found, create dynamic instance with user-defined distribution
                PAF = namedtuple('PAF', ['prod_cod', 'precio', 'moneda', 'g5', 'g10', 'exenta', 'volumen', 'peso', 'medidaobj', 'porcentaje_iva'])
                MOBJ = namedtuple('MOBJ', ['medida_cod', 'medida'])
                IOBJ = namedtuple('IOBJ', ['porcentaje'])

                exenta_pct = float(d.get('exenta', 0))
                g5_pct = float(d.get('g5', 0))
                g10_pct = float(d.get('g10', 100))

                # Determine IVA percentage based on distribution
                if g10_pct > 0:
                    porcentaje_iva_val = 10
                elif g5_pct > 0:
                    porcentaje_iva_val = 5
                else:
                    porcentaje_iva_val = 0

                medidaobj = MOBJ(medida_cod=77, medida='UNI')
                porcentajeobj = IOBJ(porcentaje=porcentaje_iva_val)

                prodobj = PAF(
                    prod_cod=prod_cod,
                    precio=precio_unitario,
                    moneda='GS',
                    g5=g5_pct,
                    g10=g10_pct,
                    exenta=exenta_pct,
                    volumen=0,
                    peso=0,
                    medidaobj=medidaobj,
                    porcentaje_iva=porcentajeobj
                )
            else:
                prod_autocreado = False

            # Use f_calcs to calculate tax amounts
            # Get discount values
            descuento_item = float(d.get('descuento', 0) or 0)
            descuento_global_item = float(d.get('descuento_global_item', 0) or 0)
            descuento_total = descuento_item + descuento_global_item

            # Calculate bruto and neto
            bruto = precio_unitario * cantidad
            neto = bruto - descuento_total
            if neto < 0:
                neto = 0

            # Calculate price based on NETO (after discount) instead of bruto
            # We need to recalculate using the effective price per unit
            if cantidad > 0 and neto > 0:
                precio_efectivo = neto / cantidad
                pcalc = f_calcs.calculate_price(prodobj, precio_efectivo, cantidad)
            else:
                pcalc = f_calcs.calculate_price(prodobj, precio_unitario, cantidad)
                # If neto is 0, all values should be 0
                if neto == 0:
                    pcalc = {
                        'exenta': 0, 'iva_5': 0, 'gravada_5': 0, 'base_gravada_5': 0,
                        'iva_10': 0, 'gravada_10': 0, 'base_gravada_10': 0
                    }

            exenta = pcalc['exenta']
            iva_5 = pcalc['iva_5']
            gravada_5 = pcalc['gravada_5']
            base_gravada_5 = pcalc['base_gravada_5']
            iva_10 = pcalc['iva_10']
            gravada_10 = pcalc['gravada_10']
            base_gravada_10 = pcalc['base_gravada_10']
            porcentaje_iva = prodobj.porcentaje_iva.porcentaje if hasattr(prodobj.porcentaje_iva, 'porcentaje') else prodobj.porcentaje_iva

            precio_unitario_siniva = precio_unitario - ((iva_10 + iva_5) / cantidad if cantidad else 0)

            # Get unidad medida code from description
            prod_unidad_medida_desc = d.get('prod_unidad_medida_desc', '77')
            try:
                prod_unidad_medida = int(prod_unidad_medida_desc)
            except:
                prod_unidad_medida = 77  # UNI default

            per_tipo_iva = 100

            if pcalc.get('exenta') and pcalc.get('afecto'):
                #total = pcalc.get('exenta')+pcalc.get('afecto')
                #per_tipo_iva = (float(prodobj.exenta)*100)/float(total)
                per_tipo_iva = prodobj.g5+prodobj.g10                

            DocumentDetail.objects.using(dbcon).create(
                documentheaderobj=docobj,
                prod_autocreado=prod_autocreado,
                prod_cod=d.get('prod_cod', 9999),
                prod_descripcion=d.get('prod_descripcion', ''),
                prod_unidad_medida=prod_unidad_medida,
                prod_unidad_medida_desc=d.get('prod_unidad_medida_desc_text', 'UNI'),
                prod_pais_origen='PRY',
                prod_pais_origen_desc='Paraguay',
                porcentaje_iva=porcentaje_iva,
                precio_unitario_source=precio_unitario,
                precio_unitario=precio_unitario,
                precio_unitario_siniva=precio_unitario_siniva,
                cantidad=cantidad,
                cantidad_devolucion=0,
                exenta_pct=Decimal(str(d.get('exenta', 0))),
                g5_pct=Decimal(str(d.get('g5', 0))),
                g10_pct=Decimal(str(d.get('g10', 100))),
                exenta=exenta,
                iva_5=iva_5,
                gravada_5=gravada_5,
                base_gravada_5=base_gravada_5,
                iva_10=iva_10,
                gravada_10=gravada_10,
                base_gravada_10=base_gravada_10,
                afecto=gravada_5 + gravada_10,
                per_tipo_iva=per_tipo_iva,
                bonifica=False,
                descuento=d.get('descuento', 0),
                per_descuento=d.get('per_descuento', 0),
                volumen=0,
                peso=0,
                observacion=None,
                cargado_usuario=userobj.first_name if userobj else 'system',
                cargado_fecha=datetime.now()
            )

        # Update totals from details
        totals = docobj.documentdetail_set.aggregate(
            exenta=Sum('exenta'),
            iva_5=Sum('iva_5'),
            gravada_5=Sum('gravada_5'),
            iva_10=Sum('iva_10'),
            gravada_10=Sum('gravada_10'),
            afecto=Sum('afecto'),
            descuento=Sum('descuento')
        )

        doc_total = (totals.get('afecto') or 0) + (totals.get('exenta') or 0)
        docobj.doc_total = doc_total
        docobj.doc_iva = (totals.get('iva_5') or 0) + (totals.get('iva_10') or 0)
        docobj.doc_exenta = totals.get('exenta') or 0
        docobj.doc_g10 = totals.get('gravada_10') or 0
        docobj.doc_i10 = totals.get('iva_10') or 0
        docobj.doc_g5 = totals.get('gravada_5') or 0
        docobj.doc_i5 = totals.get('iva_5') or 0
        docobj.doc_descuento = totals.get('descuento', 0) or 0
        # Calcular porcentaje de descuento total
        if doc_total > 0 and docobj.doc_descuento > 0:
            docobj.doc_per_descuento = (float(docobj.doc_descuento) / float(doc_total)) * 100
        else:
            docobj.doc_per_descuento = 0
        docobj.doc_redondeo = doc_redondeo

        if docobj.doc_cre_tipo_cod == 2:
            docobj.doc_saldo = docobj.get_total_venta_gs()
            docobj.doc_cre_cond = 1
            # Asegurar que ambas fechas sean objetos date antes de calcular
            vencimiento = docobj.doc_vencimiento
            if isinstance(vencimiento, str):
                vencimiento = datetime.strptime(vencimiento, '%Y-%m-%d').date()
            fecha = docobj.doc_fecha
            if isinstance(fecha, str):
                fecha = datetime.strptime(fecha, '%Y-%m-%d').date()
            docobj.doc_cre_plazo = '{} dias'.format((vencimiento - fecha).days)
            docobj.doc_cre_cond_desc = fl_sifen_conf.K_CRE_COND.get(docobj.doc_cre_cond) if docobj.doc_cre_cond else None

        docobj.doc_total_redondeo = docobj.get_total_operacion_redondeo()
        if docobj.doc_tipo == 'FE':
            docobj.doc_relacion_saldo = docobj.get_total_venta_gs()

        # For NC/ND, populate doc_loop_link with related invoice prof_number
        if docobj.doc_tipo in ['NC', 'ND'] and uc_fields.get('doc_relacion_cdc'):
            related_doc = DocumentHeader.objects.filter(
                ek_cdc=uc_fields.get('doc_relacion_cdc')
            ).first()
            if related_doc:
                docobj.doc_loop_link = related_doc.prof_number
                docobj.save()
                related_doc = DocumentHeader.objects.get(prof_number=docobj.doc_loop_link)
                total_venta_gs = sum([ x.get_total_venta_gs() for x in DocumentHeader.objects.filter(doc_loop_link=related_doc.prof_number) ])
                if total_venta_gs > related_doc.get_total_venta_gs():
                    docobj.delete()
                    return {'error': 'El valor de la NC {:,.2f} es superior al saldo de la factura {:,.2f}'.format(total_venta_gs, related_doc.doc_relacion_saldo) }, args, kwargs
                related_doc.doc_relacion_saldo = related_doc.get_total_venta_gs() - total_venta_gs
                related_doc.save()
        docobj.save()
        clobj = Clientes.objects.filter(pdv_ruc=uc_fields['pdv_ruc']).first()
        if not clobj:
            # Create new Cliente
            Clientes.objects.create(
                pdv_ruc=uc_fields['pdv_ruc'],
                pdv_ruc_dv=uc_fields['pdv_ruc_dv'],
                pdv_nombrefactura=uc_fields['pdv_nombrefactura'],
                pdv_nombrefantasia=docobj.pdv_nombrefantasia,
                pdv_celular=uc_fields['pdv_celular'] or '',
                pdv_email=uc_fields['pdv_email'] or '',
                pdv_type_business=uc_fields['pdv_type_business'],
                pdv_tipocontribuyente=uc_fields.get('pdv_tipocontribuyente'),
                pdv_es_contribuyente=uc_fields.get('pdv_es_contribuyente', True)
            )
        else:
            # Update if fields changed
            
            updated = False
            if clobj.pdv_ruc_dv != uc_fields['pdv_ruc_dv']:
                clobj.pdv_ruc_dv = uc_fields['pdv_ruc_dv']
                updated = True
            if clobj.pdv_nombrefactura != uc_fields['pdv_nombrefactura']:
                clobj.pdv_nombrefactura = uc_fields['pdv_nombrefactura']
                updated = True
            if uc_fields['pdv_celular'] and clobj.pdv_celular != uc_fields['pdv_celular']:
                clobj.pdv_celular = uc_fields['pdv_celular']
                updated = True
            if uc_fields['pdv_email'] and clobj.pdv_email != uc_fields['pdv_email']:
                clobj.pdv_email = uc_fields['pdv_email']
                updated = True
            if clobj.pdv_type_business != uc_fields['pdv_type_business']:
                clobj.pdv_type_business = uc_fields['pdv_type_business']
                updated = True
            if uc_fields.get('pdv_tipocontribuyente') and clobj.pdv_tipocontribuyente != uc_fields['pdv_tipocontribuyente']:
                clobj.pdv_tipocontribuyente = uc_fields['pdv_tipocontribuyente']
            if clobj.pdv_tipocontribuyente in [1, 2]:
                clobj.pdv_es_contribuyente = True
            else:
                clobj.pdv_es_contribuyente = False
            updated = True
            # if clobj.pdv_es_contribuyente != uc_fields.get('pdv_es_contribuyente', True):
            #     clobj.pdv_es_contribuyente = uc_fields.get('pdv_es_contribuyente', True)
            #     updated = True
            if updated:
                clobj.save()
        
        if send_save:
            rsp = self.send_to_sifen(
                userobj=userobj,
                qdict={
                    'dbcon': dbcon,
                    'pk': docobj.pk,
                    'rtype': 'update' if pk else 'create',
                }
            )
            return rsp, args, kwargs
        else:
            if not pk or docobj.doc_numero == 0 or docobj.doc_numero is None:
                rsp = self.set_number(
                    timbradoobj.timbrado,
                    [docobj.prof_number],
                    1,
                    docobj.doc_establecimiento,
                    sign_document=True,
                    doc_tipo=docobj.doc_tipo
                )
                if not rsp.get('success'): return rsp, args, kwargs
        msg = 'Documento creado exitosamente'
        if pk:
            docobj.ek_xml_ekua = False
            docobj.save()
            eser = ekuatia_serials.Eserial()
            aorde = QueryDict(mutable=True)
            aorde.update({
                'prof_number': docobj.prof_number,
                'ruc': docobj.ek_bs_ruc,
            })
            sde = eser.set_data_ekuatia(qdict=aorde)
            logging.info(f'SET DATA EKUATIA RESP: {sde}')
            msg = 'Documento actualizado exitosamente'
        return {'success': msg, 'record_id': docobj.id}, args, kwargs

    def send_to_sifen(self, *args, **kwargs) -> tuple:
        # Assign invoice number and sign
        userobj = kwargs.get('userobj')
        q: dict = kwargs.get('qdict', {})
        dbcon = q.get('dbcon', 'default')
        pk = q.get('pk')
        rtype = q.get('rtype')
        docobj = DocumentHeader.objects.using(dbcon).get(pk=pk)
        timbradoobj = Etimbrado.objects.get(timbrado=docobj.ek_timbrado)
        if rtype == 'create':
            rsp = self.set_number(
                timbradoobj.timbrado,
                [docobj.prof_number],
                1,
                docobj.doc_establecimiento,
                sign_document=True,
                doc_tipo=docobj.doc_tipo
            )
            if not rsp.get('success'): return rsp
        if rtype == 'update' and not docobj.doc_numero:
            rsp = self.set_number(
                timbradoobj.timbrado,
                [docobj.prof_number],
                1,
                docobj.doc_establecimiento,
                sign_document=True,
                doc_tipo=docobj.doc_tipo
            )
            if not rsp.get('success'): return rsp
        #send file to the sifen.
        if not docobj.ek_estado:
            #Por si es que es tipo create y se ha asignado numero.
            docobj = DocumentHeader.objects.using(dbcon).get(pk=pk)
            docobj.ek_xml_ekua = False
            docobj.save()
            eser = ekuatia_serials.Eserial()
            qek = QueryDict(mutable=True)
            qek.update({
                'prof_number': docobj.prof_number,
                'ruc': docobj.ek_bs_ruc,
            })
            eser.set_data_ekuatia(qdict=qek)
            eser.send_pending_signedxml([docobj.prof_number])
        if docobj.pdv_email is not None and docobj.pdv_email.strip():
            send_task('Sifen.tasks.send_invoice',
                kwargs={
                    'username': userobj.username,
                    'qdict': {
                        'dbcon': dbcon,
                        'docpk': docobj.id,
                        'from_console': False
                    }
                }
            )
        return {'success': 'Documento enviado a SIFEN exitosamente', 'record_id': docobj.id}

    def delete_documentheader(self, *args, **kwargs) -> dict:
        """Delete one or more DocumentHeader records - CRUD interface"""
        ios = IoS()
        q: dict = kwargs.get('qdict', {})
        dbcon = q.get('dbcon', 'default')
        ids = from_json(q.get('ids'))
        msgs = []
        if not ids:
            msgs.append({'error': 'Falta los IDs de los registros a eliminar'})
            return {'msgs': msgs}
        for pk in ids:
            mobj = DocumentHeader.objects.using(dbcon).get(pk=pk)
            if mobj.ek_estado == 'Aprobado' or mobj.lote_msg == 'Aprobado' or mobj.ek_estado == 'Rechazado' or mobj.lote_estado in ['RECIBIDO', 'PROCESANDO', 'CONCLUIDO']:
                msgs.append({'error': f'No se puede eliminar el documento {mobj.doc_numero} ya proccesado en la SIFEN'})
                continue
            if mobj.doc_tipo in ['NC', 'ND'] and mobj.doc_relacion_cdc:
                related_doc = DocumentHeader.objects.filter(
                    prof_number=mobj.doc_loop_link
                ).first()
                if related_doc:
                    related_doc.doc_relacion_saldo += mobj.get_total_venta_gs()
                    related_doc.save()
            rsp = self.release_number(
                qdict={
                    'ruc': mobj.ek_bs_ruc,
                    'establecimiento': mobj.doc_establecimiento,
                    'tipo': mobj.doc_tipo,
                    'docpk': mobj.pk,
                }
            )
            mobj.delete()
            msgs.append({'success': 'Registro eliminado exitosamente'})
            if rsp.get('error'):
                msgs.append({'error': rsp.get('error')})
            if rsp.get('success'):
                msgs.append({'success': rsp.get('success')})
        return {'msgs': msgs}

    def mostrar_xml_firmado(self, *args, **kwargs) -> dict:
        """Mostrar el contenido del XML firmado de un DocumentHeader"""
        q: dict = kwargs.get('qdict', {})
        pk = q.get('id')
        dbcon = q.get('dbcon', 'default')

        if not pk:
            return {'error': 'Falta el ID del documento'}

        mobj = DocumentHeader.objects.using(dbcon).get(pk=pk)
        if not mobj.ek_xml_file_signed:
            return {'error': 'El documento no tiene XML firmado'}

        xml_path = mobj.ek_xml_file_signed.path
        if not os.path.exists(xml_path):
            return {'error': f'El archivo XML no existe: {xml_path}'}

        with open(xml_path, 'r', encoding='utf-8') as f:
            xml_content = f.read()

        return {
            'xml_content': xml_content,
            'xml_url': mobj.ek_xml_file_signed.url,
            'xml_filename': os.path.basename(xml_path)
        }
        

    def reinicializar_lote(self, *args, **kwargs) -> dict:
        """Reinicializar el estado del lote cuando está en RECIBIDO pero sin número de lote"""
        q: dict = kwargs.get('qdict', {})
        pk = q.get('id')
        dbcon = q.get('dbcon', 'default')

        if not pk:
            return {'error': 'Falta el ID del documento'}

        
        mobj = DocumentHeader.objects.using(dbcon).get(pk=pk)

        # Verificar que el estado sea RECIBIDO y no tenga lote asignado
        if mobj.lote_estado != 'RECIBIDO':
            return {'error': f'El estado del lote debe ser RECIBIDO. Estado actual: {mobj.lote_estado}'}

        if mobj.lote and mobj.lote != '' and mobj.lote != '0':
            return {'error': f'El documento ya tiene un lote asignado: {mobj.lote}'}

        # Reinicializar el estado del lote
        mobj.lote_estado = None
        mobj.lote = 0
        mobj.lote_msg = None
        mobj.save()

        return {'success': f'Lote reinicializado para documento {mobj.doc_numero}. Ahora puede ser reenviado a SIFEN.'}
        

    def get_documentdetails(self, *args, **kwargs) -> dict:
        """Get details for a DocumentHeader - used when editing"""
        ios = IoS()
        q: dict = kwargs.get('qdict', {})
        header_id = q.get('header_id')
        dbcon = q.get('dbcon', 'default')

        if not header_id:
            return {'error': 'Falta el header_id'}

        try:
            docobj = DocumentHeader.objects.using(dbcon).get(pk=header_id)
            details = []
            for detail in docobj.documentdetail_set.filter(anulado=False):
                # Determine if product is from model or manual entry
                is_from_model = detail.prod_cod not in [9999, 1850] and not detail.prod_autocreado

                details.append({
                    'prod_cod': detail.prod_cod,
                    'prod_unidad_medida_desc': detail.prod_unidad_medida,
                    'prod_unidad_medida_desc_text': detail.prod_unidad_medida_desc,
                    'prod_descripcion': detail.prod_descripcion,
                    'precio_unitario': float(detail.precio_unitario),
                    'cantidad': float(detail.cantidad),
                    'porcentaje_iva': detail.porcentaje_iva,
                    'is_from_model': is_from_model,
                    # Percentages from database
                    'exenta_pct': float(detail.exenta_pct),
                    'g5_pct': float(detail.g5_pct),
                    'g10_pct': float(detail.g10_pct),
                    # Calculated values from database
                    'exenta': float(detail.exenta),
                    'iva_5': float(detail.iva_5),
                    'gravada_5': float(detail.gravada_5),
                    'base_gravada_5': float(detail.base_gravada_5),
                    'iva_10': float(detail.iva_10),
                    'gravada_10': float(detail.gravada_10),
                    'base_gravada_10': float(detail.base_gravada_10),
                    'subtotal': float(detail.exenta) + float(detail.gravada_5) + float(detail.gravada_10),
                    # Discount values
                    'descuento': float(detail.descuento or 0),
                    'per_descuento': float(detail.per_descuento or 0)
                })
            # Calculate global discount percentage from header
            total_bruto = sum(d['precio_unitario'] * d['cantidad'] for d in details)
            doc_descuento_global = float(docobj.doc_descuento_global or 0)
            doc_descuento_global_pct = (doc_descuento_global / total_bruto * 100) if total_bruto > 0 else 0

            return {
                'success': True,
                'details': details,
                'doc_descuento_global': doc_descuento_global,
                'doc_descuento_global_pct': doc_descuento_global_pct
            }
        except DocumentHeader.DoesNotExist:
            return {'error': 'Documento no encontrado'}

    def validate_ruc(self, *args, **kwargs) -> dict:
        """Validate RUC number"""
        ios = IoS()
        q: dict = kwargs.get('qdict', {})
        ruc = q.get('ruc', '').strip()
        eser = ekuatia_serials.Eserial()
        clobj = Clientes.objects.filter(pdv_ruc=ruc).first()
        if clobj:
            if clobj.pdv_es_contribuyente == False:
                rsp = eser.qr_ruc(ruc, business=self.bsobj)
                if rsp.get('dmsgres') == 'RUC encontrado':
                    clobj.pdv_es_contribuyente = True
                    clobj.pdv_innominado = False
                    clobj.save()
            return {
                'success': 'RUC válido',
                'pdv_ruc_dv': clobj.pdv_ruc_dv,
                'pdv_nombrefactura': clobj.pdv_nombrefactura,
                'pdv_celular': clobj.pdv_celular,
                'pdv_email': clobj.pdv_email,
                'pdv_type_business': clobj.pdv_type_business,
                'pdv_tipocontribuyente': clobj.pdv_tipocontribuyente,
            }
        rsp = eser.qr_ruc(ruc, business=self.bsobj)
        if rsp.get('dmsgres') != 'RUC encontrado':
            return {'error': 'RUC no válido o no encontrado'}
        gdata = mng_gmdata.Gdata()
        pdv_ruc_dv = gdata.calculate_dv(ruc)
        return {
            'success': 'RUC válido',
            'pdv_ruc_dv': pdv_ruc_dv,
            'pdv_nombrefactura': rsp.get('drazcons').strip(),
            'pdv_celular': '',
            'pdv_email': '',
        }

    def trace_lote(self, *args, **kwargs) -> dict:
        """Wrapper method to call qr_lote from SoapSifen"""
        from Sifen.rq_soap_handler import SoapSifen
        from Sifen.models import SoapMsg
        q: dict = kwargs.get('qdict', {})
        soapmsg_id = q.get('soapmsg_id')
        dbcon = q.get('dbcon', 'default')

        soapobj = SoapMsg.objects.using(dbcon).get(pk=soapmsg_id)
        lote = soapobj.json_rsp.get('dprotconslote')

        soap_sifen = SoapSifen(business=self.bsobj)
        rsp = soap_sifen.qr_lote(lote)

        return {'success': f'Consulta de lote {lote} ejecutada', 'response': str(rsp.text)}

    def create_cliente(self, *args, **kwargs) -> tuple:
        """Crea o actualiza un registro de Cliente"""
        ios = IoS()
        userobj = kwargs.get('userobj')
        q: dict = kwargs.get('qdict', {})
        dbcon = q.get('dbcon', 'default')
        uc_fields: dict = from_json(q.get('uc_fields', {}))
        rnorm, rrm, rbol = ios.format_data_for_db(Clientes, uc_fields)
        for c in rrm: uc_fields.pop(c)
        for f, fv in rbol: uc_fields[f] = fv
        for f, fv in rnorm: uc_fields[f] = fv
        ff = ios.form_model_fields(uc_fields, Clientes._meta.fields)
        for rr in ff:
            uc_fields.pop(rr)
        pk = uc_fields.get('id')
        msg = 'Cliente creado exitosamente'
        files: dict = kwargs.get('files')
        if pk:
            mobj = Clientes.objects.get(pk=pk)
            msg = 'Cliente actualizado exitosamente'
            u_fields = ios.get_differences_fields(model_to_dict(mobj), uc_fields)
            if not u_fields and not files:
                return {'info': 'Nada que actualizar'}, args, kwargs
            if not u_fields and files:
                msg = 'Archivos actualizados exitosamente'
            if u_fields:
                Clientes.objects.using(dbcon).filter(pk=mobj.pk).update(**u_fields)
            if u_fields and files:
                msg = 'Cliente y archivos actualizados exitosamente'
        else:
            uc_fields['cargado_usuario'] = userobj.username if userobj else None
            mobj = Clientes.objects.using(dbcon).create(**uc_fields)
        return {'success': msg, 'record_id': mobj.id}, args, kwargs

    def delete_cliente(self, *args, **kwargs) -> dict:
        """Elimina uno o mas registros de Cliente"""
        ios = IoS()
        q: dict = kwargs.get('qdict', {})
        dbcon = q.get('dbcon', 'default')
        ids = from_json(q.get('ids'))
        msgs = []
        if not ids:
            msgs.append({'error': 'Falta los IDs de los registros a eliminar'})
            return {'msgs': msgs}
        for pk in ids:
            mobj = Clientes.objects.using(dbcon).get(pk=pk)
            mobj.delete()
            msgs.append({'success': 'Cliente eliminado exitosamente'})
        return {'msgs': msgs}
